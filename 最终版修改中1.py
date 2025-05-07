# -*- coding: utf-8 -*-

# 视频用户行为模拟系统 - 综合版
# 功能特点：
# 1. 从Excel加载视频数据（包含URL和关键词）-> 修改为：先将Excel转为CSV，再从CSV加载
# 2. 使用TF-IDF和层次聚类对视频自动分类（最多400类）
# 3. 生成模拟用户及其兴趣偏好
# 4. 模拟7天的用户观看和点赞行为
# 5. 生成详细报告，包含6个工作表
#   - 分类统计
#   - 观看记录（按天分列）
#   - 点赞记录
#   - 视频统计数据
#   - 每日观看统计
#   - 热门视频排行
# 6. 给视频进行聚类分析，将具有相似观看用户的视频聚成一团
# 7. 给用户进行聚类分析，将具有相似观看兴趣的用户聚成一团
# 8. 支持从之前报告中恢复分类信息

# ==================== 导入依赖库 ====================
import random
import os
import sys # Import sys for potentially exiting on critical errors
from datetime import datetime, timedelta
from collections import defaultdict, Counter
import json
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.cluster import DBSCAN
from sklearn.preprocessing import StandardScaler
# Make sure openpyxl is installed if you intend to use convert_excel_to_csv
try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("Warning: openpyxl library not found. Excel operations (like conversion) might fail.")
    Workbook = None # Assign None to avoid errors if not used directly later

from sklearn.feature_extraction.text import TfidfVectorizer, HashingVectorizer
from sklearn.cluster import AgglomerativeClustering, MiniBatchKMeans
import numpy as np
import traceback
from tqdm import tqdm
import pandas as pd
from sklearn.decomposition import TruncatedSVD
from sklearn.pipeline import Pipeline
from sklearn.metrics.pairwise import cosine_similarity
import gc
import time
from multiprocessing import Pool, cpu_count
from sklearn.cluster import AgglomerativeClustering, MiniBatchKMeans, Birch # Added BIRCH
# 导入新功能模块 (Ensure these files exist in the same directory or Python path)
try:
    import video_clustering
    import user_clustering
    import category_recovery
    import user_center
    import videos_playing
except ImportError as e:
    print(f"Error importing custom modules: {e}")
    print("Please ensure video_clustering.py, user_clustering.py, and category_recovery.py are present.")
    # sys.exit(1) # Optional: Exit if modules are critical

# ==================== 全局配置 ====================
NUM_USERS = 10000
MIN_INTERESTS = 4
MAX_INTERESTS = 6
OUTPUT_FILENAME = "user_behavior_7days.xlsx"
DAYS_TO_SIMULATE = 7
# Define source Excel and target CSV paths clearly
SOURCE_EXCEL_PATH = "D:/Desktop/数据结构大作业/merge(4).xlsx"
TARGET_CSV_PATH = "D:/Desktop/数据结构大作业/merge(4).csv"


def clean_excel_data(text):
    if text is None:
        return ""

    #转换为字符串
    text = str(text)
    #移除所有控制字符（ASCII 0-31，除了\t,\n,\r）
    text = "".join(ch for ch in text if ord(ch) >= 32 or ch in ("\t", "\n", "\r"))
    #替换特定特殊字符
    replacements = {
        "|": "-",
        "\t": " ",  # 制表符替换为空格
        "\n": " ",  # 换行符替换为空格
        "\r": " ",  # 回车符替换为空格
        "\x0b": " ",  # 垂直制表符
        "\x1c": " ",  # 文件分隔符
        "\x1d": " ",  # 组分隔符
        "\x1e": " ",  # 记录分隔符
        "\x1f": " ",  # 单元分隔符
        "": "",  # 直接移除这个特殊字符
        "": "",  # 移除其他可能存在的控制字符
        "": "",
        "": "",
        "": ""
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    #移除不可打印的Unicode字符（如零宽空格等）
    text = "".join(ch for ch in text if ch.isprintable() or ch == " ")
    #标准化空格（多个空格变单个）
    text = " ".join(text.split())
    #截断过长的文本（Excel单元格限制约32,767个字符）
    return text[:32000] if len(text) > 32000 else text


def convert_excel_to_csv(excel_path, csv_path):
    """将Excel文件转换为CSV格式 (分批处理，优化内存)"""
    try:
        import pandas as pd
        print(f"开始将 '{os.path.basename(excel_path)}' 转换为CSV...")
        # 直接使用单次读取方式 (不使用chunksize参数)
        print("读取Excel文件 (可能需要一些时间)...")
        df = pd.read_excel(excel_path, engine='openpyxl')
        total_converted = len(df)
        # 分批写入CSV以减少内存使用
        print(f"总共 {total_converted} 行数据，开始分批写入CSV...")
        chunk_size = 50000
        with open(csv_path, 'w', encoding='utf-8', newline='') as f:
            # 写入标题行
            df.head(0).to_csv(f, index=False, header=True)
            # 分批写入数据
            for i in range(0, len(df), chunk_size):
                chunk_end = min(i + chunk_size, len(df))
                chunk = df.iloc[i:chunk_end]
                chunk.to_csv(f, index=False, header=False, mode='a')
                print(f"已写入 {chunk_end} / {total_converted} 行...")
                # 释放内存
                del chunk
                gc.collect()
        # 释放内存
        del df
        gc.collect()
        print(f"转换完成！共 {total_converted} 行数据。CSV文件保存于：{csv_path}")
        return True
    except ImportError:
        print("错误：需要安装 'pandas' 和 'openpyxl' 库才能转换Excel。请运行 'pip install pandas openpyxl'")
        return False
    except FileNotFoundError:
        print(f"错误：无法找到源Excel文件 '{excel_path}'")
        return False
    except Exception as e:
        print(f"Excel转换为CSV时发生错误：{str(e)}")
        # 清理可能不完整的CSV文件
        if os.path.exists(csv_path):
            try:
                os.remove(csv_path)
                print(f"已删除不完整的CSV文件: {csv_path}")
            except OSError as rm_err:
                print(f"警告：无法删除不完整的CSV文件 '{csv_path}': {rm_err}")
        return False


# ==================== 核心类定义 ====================
class Video:
    """优化内存的视频类"""
    __slots__ = ['id','url', 'keywords', 'category', 'title', 'like_count',
                 'cover_url', 'play_url', 'viewers', 'liked_users']

    def __init__(self, video_id,url, keywords, title=None, cover_url=None, play_url=None): # 添加了 title 参数
        self.id = str(video_id).strip() if video_id is not None else "N/A_ID"  # 确保是字符串并处理 None
        self.url = url.strip() if url else "N/A"  # 处理可能为 None 的 URL
        # 优雅地处理可能为 NaN 或 None 的关键词
        keywords_str = str(keywords) if keywords is not None else ""
        self.keywords = tuple(keywords_str.split()[:15])  # 限制关键词数量，处理空字符串
        self.category = None
        self.title = str(title).strip() if title else None  # 存储标题
        self.like_count = 0
        self.cover_url = cover_url.strip()[:200] if cover_url and pd.notna(cover_url) else None  # 限制 URL 长度，处理 NaN
        self.play_url = play_url.strip()[:200] if play_url and pd.notna(play_url) else None  # 限制 URL 长度，处理 NaN
        self.viewers = set()  # 使用集合来存储独立的观看者
        self.liked_users = set()  # 使用集合来存储独立的点赞用户



class User:
    """用户对象，包含兴趣偏好和行为记录"""

    def __init__(self, user_id, all_categories):
        self.user_id = f"U{user_id:04d}"
        self.interests = self._generate_interests(all_categories)
        self.watched_videos = []  # Format: [(url, timestamp_str), ...]
        self.liked_videos = []  # Format: [(url, timestamp_str), ...]

    def _generate_interests(self, categories):
        """随机生成用户兴趣分布"""
        if not categories:
            # print("警告: 没有有效的类别传入，用户兴趣将使用默认值。")
            return {'默认类别': 1.0} # Provide default interest if no valid categories exist
        valid_categories = list(categories) # Ensure it's a list
        if not valid_categories: # Double check after potential filtering
             return {'默认类别': 1.0}

        # Ensure num_interests does not exceed the number of available categories
        max_possible_interests = len(valid_categories)
        min_req_interests = min(MIN_INTERESTS, max_possible_interests)
        max_req_interests = min(MAX_INTERESTS, max_possible_interests)

        # Handle edge case where MAX_INTERESTS < MIN_INTERESTS or few categories
        if min_req_interests > max_req_interests:
             num_interests = max_possible_interests # Assign all available if range is invalid
        else:
             num_interests = random.randint(min_req_interests, max_req_interests)

        # Ensure sampling size doesn't exceed population size
        num_interests = min(num_interests, len(valid_categories))

        if num_interests == 0:
             return {'默认类别': 1.0} # Fallback if somehow num_interests becomes 0

        chosen_cats = random.sample(valid_categories, num_interests)
        return {cat: random.uniform(0.7, 1.0) for cat in chosen_cats}


# --- 修改 load_csv_data 函数 ---
def load_csv_data(file_path):
    """从CSV加载视频数据（支持大数据量，优化内存），包含视频ID (列名为 'id')"""
    try:
        import pandas as pd # 保留 import
        print(f"开始从 '{os.path.basename(file_path)}' 加载视频数据 (ID列名为 'id')...")

        # --- 修改：将 'video_id' 改为 'id' ---
        required_cols = ['id', 'video_url', 'source_keyword']
        optional_cols = ['title', 'video_cover_url', 'video_play_url']
        # use_cols 会自动更新，因为它依赖 required_cols
        use_cols = required_cols + optional_cols

        # 检查文件头
        try:
            header = pd.read_csv(file_path, nrows=0, encoding='utf-8').columns.tolist()
            missing_required = [col for col in required_cols if col not in header]
            if missing_required:
                # --- 修改：更新了错误信息中的列名 ---
                raise ValueError(f"CSV文件 '{file_path}' 缺少必需的列: {', '.join(missing_required)}。请确保包含 'id', 'video_url', 'source_keyword' 列。")
            actual_use_cols = [col for col in use_cols if col in header]
            print(f"将加载以下列: {', '.join(actual_use_cols)}")
        except Exception as e:
             raise ValueError(f"无法读取CSV文件头 '{file_path}': {e}")

        # 配置数据类型
        # --- 修改：将 dtype_config 中的键 'video_id' 改为 'id' ---
        dtype_config = {
            'id': 'object', # CSV 列名为 'id'
            'video_url': 'object',
            'source_keyword': 'object',
            'title': 'object',
            'video_cover_url': 'object',
            'video_play_url': 'object'
        }
        effective_dtype = {k: v for k, v in dtype_config.items() if k in actual_use_cols}

        # 估算总行数 (逻辑不变)
        total_rows = 0
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                total_rows = sum(1 for row in f) -1
            if total_rows < 0: total_rows = 0
        except Exception as e:
            print(f"警告：无法准确计算行数，进度条可能不准确。错误: {e}")
            total_rows = None

        videos = []
        seen_urls = set()
        seen_ids = set() # Optional: Track seen IDs as well
        processed_rows = 0
        chunksize = 10000

        pbar = tqdm(total=total_rows, desc="加载CSV数据", unit="rows", disable=(total_rows is None))

        for chunk in pd.read_csv(
                file_path,
                chunksize=chunksize,
                dtype=effective_dtype,
                usecols=actual_use_cols,
                encoding='utf-8',
                on_bad_lines='warn'
        ):
            # 数据清理
            # --- 修改：在 dropna 的 subset 中将 'video_id' 改为 'id' ---
            chunk.dropna(subset=['id', 'video_url', 'source_keyword'], inplace=True)
            # --- 修改：更新关于基于 ID 去重的注释 (如果需要) ---
            # 当前仍基于 URL 去重。
            chunk.drop_duplicates(subset='video_url', keep='first', inplace=True)
            # 如果需要基于 'id' 列去重，使用下面这行:
            # chunk.drop_duplicates(subset='id', keep='first', inplace=True)

            # 转换为 Video 对象
            for _, row in chunk.iterrows():
                url = str(row['video_url']).strip()
                # --- 修改：从 row['id'] 获取 ID 值 ---
                video_id_from_csv = str(row['id']).strip() # 从 'id' 列读取

                # 基于 URL 的唯一性检查 (保持不变)
                if url and url not in seen_urls:
                    seen_urls.add(url)
                    # Optional: Add ID check using video_id_from_csv
                    # if video_id_from_csv in seen_ids:
                    #     print(f"警告：发现重复的 ID '{video_id_from_csv}' 对应不同的 URL '{url}'. 已跳过.")
                    #     continue
                    # seen_ids.add(video_id_from_csv)

                    # --- 修改：将从 'id' 列读取的值传递给 Video 构造函数的 video_id 参数 ---
                    videos.append(Video(
                        video_id=video_id_from_csv, # 将 CSV 的 'id' 列的值传给构造函数的 video_id 参数
                        url=url,
                        keywords=row['source_keyword'],
                        title=row.get('title'),
                        cover_url=row.get('video_cover_url'),
                        play_url=row.get('video_play_url')
                    ))

            processed_rows += len(chunk)
            if total_rows is not None:
                pbar.update(len(chunk))
            else:
                pbar.set_description(f"加载CSV数据 (已处理 {processed_rows} 行)")
            gc.collect()

        pbar.close()
        if not videos:
             # --- 修改：更新了警告信息中的列名 ---
             print(f"\n警告：从 '{file_path}' 加载了 0 个有效视频。请检查CSV文件内容、格式和必需列 ('id', 'video_url', 'source_keyword')。")
        else:
             # 成功信息保持不变，因为它描述的是概念上的 ID
             print(f"\n成功加载 {len(videos)} 个唯一视频（含视频ID，内存优化版）")
        return videos

    except FileNotFoundError:
        raise FileNotFoundError(f"错误：CSV 数据文件未找到 '{file_path}'")
    except ValueError as ve:
         raise ve
    except Exception as e:
        raise RuntimeError(f"加载 CSV 数据时发生意外错误: {e.__class__.__name__} - {str(e)}") from e



from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import TfidfVectorizer
import numpy as np


# ==================== 数据处理函数 ====================
# ... (other functions like load_csv_data remain the same) ...

def auto_categorize(videos, n_categories=400):
    """优化版视频分类：支持10万级数据，强制400类，均匀分布"""
    print(f"\n开始自动分类 (10万级数据优化版，强制{n_categories}类)...")

    # ===== 第一阶段：数据预处理 =====
    print("1. 数据预处理...")
    video_docs = []
    valid_indices = []

    # 使用生成器表达式减少内存占用
    for idx, v in enumerate(videos):
        if v.keywords:
            doc = ' '.join(v.keywords[:15])  # 限制关键词数量
            if doc.strip():
                video_docs.append(doc)
                valid_indices.append(idx)

    total_videos = len(video_docs)
    if total_videos == 0:
        print("警告：没有有效关键词的视频")
        return {}

    # ===== 第二阶段：特征工程 =====
    print("2. 特征提取 (使用HashingVectorizer避免内存爆炸)...")
    vectorizer = HashingVectorizer(
        n_features=2 ** 18,  # 增大特征空间
        alternate_sign=False,  # 提高稀疏矩阵效率
        ngram_range=(1, 2),  # 包含二元语法
        dtype=np.float32  # 减少内存占用
    )
    X = vectorizer.fit_transform(video_docs)

    # ===== 第三阶段：分层聚类 =====
    print("3. 分层聚类 (处理大规模数据)...")

    # 计算动态类别容量范围
    avg_size = total_videos / n_categories
    max_size = int(avg_size * 1.8)  # 上限放宽
    min_size = int(avg_size * 0.7)  # 下限放宽

    # 使用MiniBatchKMeans进行初始粗分类
    print(" - 初始粗分类 (50个簇)...")
    mbk = MiniBatchKMeans(
        n_clusters=50,
        batch_size=5000,
        compute_labels=True,
        random_state=42
    )
    coarse_labels = mbk.fit_predict(X)

    # ===== 第四阶段：动态细分类 =====
    print("4. 动态细分类 (确保400类)...")
    final_labels = np.zeros(total_videos, dtype=np.int16)
    current_cat = 0
    category_sizes = np.zeros(n_categories, dtype=np.int32)

    # 进度条
    pbar = tqdm(total=50, desc="处理粗分类簇")

    for cluster_id in range(50):
        # 获取当前粗分类中的视频索引
        cluster_mask = (coarse_labels == cluster_id)
        cluster_indices = np.where(cluster_mask)[0]
        cluster_size = len(cluster_indices)

        if cluster_size == 0:
            pbar.update(1)
            continue

        # 计算需要划分的子类数量
        n_sub_clusters = max(1, min(
            n_categories - current_cat,
            int(np.ceil(cluster_size / avg_size))
        ))

        if n_sub_clusters == 1:
            # 直接分配类别（带容量检查）
            for idx in cluster_indices:
                if category_sizes[current_cat] < max_size:
                    final_labels[idx] = current_cat
                    category_sizes[current_cat] += 1
                else:
                    current_cat = min(current_cat + 1, n_categories - 1)
                    final_labels[idx] = current_cat
                    category_sizes[current_cat] += 1
        else:
            # 使用SVD降维后K-means
            print(f" - 簇{cluster_id} 细分为{n_sub_clusters}类...")
            sub_X = X[cluster_indices]

            # 使用TruncatedSVD降维
            svd = TruncatedSVD(n_components=100)
            reduced_X = svd.fit_transform(sub_X)

            # 细分类
            sub_kmeans = MiniBatchKMeans(
                n_clusters=n_sub_clusters,
                batch_size=2000,
                random_state=42
            )
            sub_labels = sub_kmeans.fit_predict(reduced_X)

            # 分配子类（带容量控制）
            for sub_id in range(n_sub_clusters):
                sub_mask = (sub_labels == sub_id)
                sub_indices = cluster_indices[sub_mask]

                for idx in sub_indices:
                    if category_sizes[current_cat] < max_size:
                        final_labels[idx] = current_cat
                        category_sizes[current_cat] += 1
                    else:
                        current_cat = min(current_cat + 1, n_categories - 1)
                        final_labels[idx] = current_cat
                        category_sizes[current_cat] += 1

        pbar.update(1)
    pbar.close()

    # ===== 第五阶段：平衡优化 =====
    print("5. 平衡优化...")

    # 找出需要调整的类别
    overfull = np.where(category_sizes > max_size)[0]
    underfull = np.where(category_sizes < min_size)[0]

    # 使用稀疏矩阵存储类别相似度
    print(" - 计算类别相似度矩阵...")
    cat_centroids = np.zeros((n_categories, X.shape[1]))
    for cat in range(n_categories):
        cat_mask = (final_labels == cat)
        if np.sum(cat_mask) > 0:
            cat_centroids[cat] = X[cat_mask].mean(axis=0)

    # 计算余弦相似度
    similarity_matrix = cosine_similarity(cat_centroids)

    # 重新分配过大的类别
    print(" - 重新分配视频...")
    for src_cat in overfull:
        # 找到最相似且有容量的目标类别
        sim_scores = similarity_matrix[src_cat]
        sorted_cats = np.argsort(-sim_scores)  # 降序排列

        for dst_cat in sorted_cats:
            if dst_cat == src_cat:
                continue
            if category_sizes[dst_cat] < max_size:
                # 计算可转移数量
                transfer_num = min(
                    category_sizes[src_cat] - max_size,
                    max_size - category_sizes[dst_cat]
                )
                if transfer_num <= 0:
                    continue

                # 转移视频（选择距离目标类别最近的）
                src_indices = np.where(final_labels == src_cat)[0]
                src_vectors = X[src_indices]

                dst_centroid = cat_centroids[dst_cat].reshape(1, -1)
                distances = cosine_similarity(src_vectors, dst_centroid).flatten()
                closest_indices = np.argsort(distances)[-transfer_num:]

                final_labels[src_indices[closest_indices]] = dst_cat
                category_sizes[src_cat] -= transfer_num
                category_sizes[dst_cat] += transfer_num
                break

    # ===== 第六阶段：结果分配 =====
    print("6. 分配最终类别...")
    category_map = {i: f"CAT_{i:03d}" for i in range(n_categories)}
    for i, idx in enumerate(valid_indices):
        videos[idx].category = category_map[final_labels[i]]

    # ===== 统计报告 =====
    print("\n分类完成！统计信息：")
    size_stats = pd.Series(category_sizes).describe()
    print(size_stats)
    print(f"最大类别大小: {max(category_sizes)}")
    print(f"最小类别大小: {min(category_sizes)}")

    return Counter(v.category for v in videos if v.category)


# ==================== 行为模拟函数 ====================
def simulate_behavior(users, videos):
    print(f"\n开始模拟 {DAYS_TO_SIMULATE} 天的用户行为...")
    if not users:
        print("警告: 没有用户数据，无法模拟行为。")
        return
    if not videos:
        print("警告: 没有视频数据，无法模拟行为。")
        return

    # Index videos by category for faster lookup
    cat_index = defaultdict(list)
    valid_videos = [] # Keep track of videos with categories for simulation
    for v in videos:
        if v.category and v.category not in ["无关键词", "其他", "分类错误", "单一类别"]: # Exclude non-specific categories
            cat_index[v.category].append(v)
            valid_videos.append(v)

    if not cat_index or not valid_videos:
        print("警告: 没有找到具有有效类别的视频。模拟可能不准确，将使用所有视频。")
        cat_index.clear() # Clear index if it's not useful
        valid_videos = list(videos) # Use all videos as fallback
        # Ensure all videos are in a default category for random selection if needed
        if valid_videos:
            cat_index['所有视频'] = valid_videos


    # Generate date range for simulation
    base_date = datetime.now().date()
    date_range = [base_date - timedelta(days=i) for i in range(DAYS_TO_SIMULATE - 1, -1, -1)]

    # Ensure each video gets at least one view (simplistic approach)
    # Convert list to set for efficient removal, then back to list if needed
    unwatched_videos = set(v for v in videos if v.url != "N/A") # Use the actual video objects


    # --- Simulation Loop ---
    total_simulated_watches = 0
    total_simulated_likes = 0

    for user in tqdm(users, desc="模拟用户行为"):
        if not user.interests or user.interests == {'默认类别': 1.0}:
            # User has no specific interests, maybe assign random popular categories?
            # For now, they will mostly explore randomly
            user_interest_categories = list(cat_index.keys()) # Use all available categories
            user_interest_weights = [1.0] * len(user_interest_categories) if user_interest_categories else []
        else:
            # Filter interests to only include categories present in the video data
            user_interest_categories = [cat for cat in user.interests.keys() if cat in cat_index]
            user_interest_weights = [user.interests[cat] for cat in user_interest_categories]


        for day in date_range:
            # Simulate 1 to 3 sessions per day
            for _ in range(random.randint(1, 2)):
                # Random time within the day (e.g., 8 AM to 11 PM)
                session_base_time = datetime.combine(day, datetime.min.time())
                session_time = session_base_time.replace(
                    hour=random.randint(8, 22),
                    minute=random.randint(0, 59),
                    second=random.randint(0, 59)
                )

                # Simulate 10 to 30 actions (watches/likes) per session
                for _ in range(random.randint(13, 15)):
                    chosen_video = None
                    # Decide whether to watch based on interest (90% chance) or explore randomly (10%)
                    if user_interest_categories and random.random() < 0.9:
                        try:
                             # Choose a category based on user's interest weights
                             chosen_cat = random.choices(user_interest_categories, weights=user_interest_weights, k=1)[0]
                             # Get candidate videos from that category
                             candidates = cat_index.get(chosen_cat, [])
                             if candidates:
                                 chosen_video = random.choice(candidates)
                        except IndexError:
                             # This might happen if weights list is empty or doesn't match categories
                             chosen_cat = None # Fallback to random exploration
                        except Exception as e:
                             print(f"Warning: Error choosing category for user {user.user_id}: {e}")
                             chosen_cat = None

                    # If no interest-based choice or exploring randomly, pick from any valid video
                    if not chosen_video:
                         if valid_videos: # Ensure there are videos to choose from
                              chosen_video = random.choice(valid_videos)

                    # Record watch if a video was chosen
                    if chosen_video:
                        _record_watch(user, chosen_video, session_time)
                        total_simulated_watches += 1
                        # Mark video as watched (remove from unwatched set)
                        unwatched_videos.discard(chosen_video) # Efficiently removes if present

                        # Simulate liking based on interest and randomness
                        like_probability = 0.1 # Base like probability
                        if chosen_video.category in user.interests:
                            # Increase probability if video matches user interest
                            like_probability += user.interests[chosen_video.category] * 0.4 # Scaled interest contribution

                        if random.random() < like_probability:
                            if _record_like(user, chosen_video, session_time): # Check if like was actually recorded
                                 total_simulated_likes +=1


    # --- Handle unwatched videos ---
    # Force watch remaining unwatched videos to ensure all have at least one view
    if unwatched_videos:
        print(f"\n强制观看 {len(unwatched_videos)} 个在模拟中未被观看的视频...")
        # Distribute these watches among users somewhat randomly
        users_list = list(users) # Convert to list if needed
        fallback_time = datetime.combine(date_range[0], datetime.min.time()).replace(hour=9) # Default time for forced watches
        for video in tqdm(list(unwatched_videos), desc="处理未观看视频"): # Convert set to list for tqdm
             if users_list: # Check if there are users to assign the watch to
                  chosen_user = random.choice(users_list)
                  _record_watch(chosen_user, video, fallback_time)
                  total_simulated_watches += 1
             else:
                  print(f"警告：没有用户可分配强制观看给视频 {video.url}")

    print(f"\n行为模拟完成。总模拟观看: {total_simulated_watches}, 总模拟点赞: {total_simulated_likes}")
    # Clean up large temporary structures if needed
    del cat_index, valid_videos, unwatched_videos
    gc.collect()


def _record_watch(user, video, timestamp):
    """Records a watch event for a user and video."""
    if not video or not video.url or video.url == "N/A":
         return # Do not record watch for invalid video object or URL

    entry = (video.url, timestamp.strftime('%Y-%m-%d %H:%M:%S'))
    user.watched_videos.append(entry)
    # Add user ID to the video's viewers set (efficiently handles duplicates)
    video.viewers.add(user.user_id)
    # No need to return anything, modifies objects in place


def _record_like(user, video, timestamp):
    """Records a like event if the user hasn't liked this video before."""
    if not video or not video.url or video.url == "N/A":
        return False # Cannot like invalid video

    # Check if user has already liked this specific video URL
    if user.user_id not in video.liked_users:
        entry = (video.url, timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        user.liked_videos.append(entry)
        video.like_count += 1
        video.liked_users.add(user.user_id) # Add user to set of likers for this video
        return True # Like was recorded
    return False # Like was not recorded (already liked)


# ==================== Daily View Statistics ====================
def count_daily_views(videos, users, days_to_simulate=DAYS_TO_SIMULATE):
    """Calculates the number of views per video for each simulated day."""
    print("\n计算每日观看统计...")
    if not videos: return {} # Return empty dict if no videos

    # Generate the date strings for the simulated period
    base_date = datetime.now().date()
    date_range_str = [(base_date - timedelta(days=i)).strftime('%Y-%m-%d')
                      for i in range(days_to_simulate - 1, -1, -1)]

    # Initialize stats dictionary: {video_url: {date_str: 0, ...}, ...}
    daily_stats = {v.url: {date_str: 0 for date_str in date_range_str}
                   for v in videos if v.url != "N/A"}

    # Aggregate views from user history
    for user in tqdm(users, desc="统计每日观看", unit="user"):
        for url, timestamp_str in user.watched_videos:
            try:
                # Extract the date part from the timestamp string
                watch_date_str = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                # Increment count if the URL and date are valid
                if url in daily_stats and watch_date_str in daily_stats[url]:
                    daily_stats[url][watch_date_str] += 1
            except ValueError:
                # Handle potential malformed timestamp strings, though unlikely with strftime
                print(f"警告：跳过格式错误的观看记录时间戳：'{timestamp_str}' for URL '{url}'")
            except KeyError:
                 # Handle cases where URL might not be in daily_stats (e.g., added after init)
                 # Or date is outside the simulated range (shouldn't happen with current logic)
                 # print(f"警告：跳过观看记录，URL '{url}' 或日期 '{watch_date_str}' 不在统计范围内。")
                 pass # Silently ignore if not critical

    print("每日观看统计计算完成。")
    return daily_stats


# ==================== Hot Video Ranking ====================
# Using Python's built-in sort (Timsort) is generally efficient and stable.
# Re-implementing merge sort isn't necessary unless for specific academic reasons.

def generate_hot_ranking(videos, top_n=100):
    """Generates a ranked list of top N hot videos based on views and likes."""
    print(f"\n生成Top {top_n} 热门视频排行...")
    if not videos:
        print("没有视频数据可供排行。")
        return []

    # Define the hotness score calculation function
    def calculate_hotness(video):
        # Weighted score: 60% views, 40% likes
        # Use len(video.viewers) for unique view count, video.like_count for total likes
        view_count = len(video.viewers)
        like_count = video.like_count # Already tracks unique likes per user implicitly
        return (view_count * 0.6) + (like_count * 0.4)

    # Sort videos by hotness score in descending order
    # Using lambda function with sort is concise and efficient
    # Add a secondary sort key (e.g., URL) for deterministic order in case of ties
    try:
        sorted_videos = sorted(
            [v for v in videos if v.url != "N/A"], # Filter out invalid videos
            key=lambda v: (calculate_hotness(v), v.url), # Primary: hotness (desc), Secondary: URL (asc)
            reverse=True # Sort by hotness descending
        )
    except Exception as e:
        print(f"错误：排序视频时出错: {e}")
        return [] # Return empty list on error


    # Build the ranking list with details for the top N videos
    ranking_list = []
    for rank, video in enumerate(sorted_videos[:top_n], 1):
        hotness_score = calculate_hotness(video)
        ranking_list.append({
            "排名": rank,
            "视频URL": video.url,
            "类别": video.category or "N/A", # Handle None category
            "观看数 (独立用户)": len(video.viewers),
            "点赞数": video.like_count,
            "热度评分": round(hotness_score, 2) # Round for display
            # Add title, cover_url, play_url if needed directly in ranking data
            # "标题": video.title or "N/A",
            # "封面地址": video.cover_url or "N/A",
            # "播放地址": video.play_url or "N/A",
        })

    print(f"热门视频排行生成完成 (Top {len(ranking_list)})。")
    return ranking_list


# ==================== Report Generation Functions ====================
def save_detailed_report(users, videos, filename=OUTPUT_FILENAME):
    """Generates and saves a multi-sheet Excel report."""
    print(f"\n开始生成详细报告 '{filename}'...")
    if Workbook is None:
        print("错误：无法生成Excel报告，因为 'openpyxl' 库未安装。")
        return

    start_time = time.time()
    try:
        wb = Workbook() # Create a new workbook

        # --- Create each sheet ---
        print("  - 创建 分类统计 sheet...")
        _create_category_sheet(wb, videos) # wb.active is sheet 1

        print("  - 创建 观看记录 sheet...")
        _create_watch_history_sheet(wb, users)

        print("  - 创建 点赞记录 sheet...")
        _create_like_history_sheet(wb, users)

        print("  - 创建 视频统计 sheet...")
        _create_video_stats_sheet(wb, videos)

        print("  - 创建 每日观看 sheet...")
        daily_views_data = count_daily_views(videos, users) # Calculate once
        _create_daily_views_sheet(wb, videos, daily_views_data)

        print("  - 创建 热门排行 sheet...")
        hot_ranking_data = generate_hot_ranking(videos) # Calculate once
        _create_hot_ranking_sheet(wb, videos, hot_ranking_data)

        # --- Save the workbook ---
        print(f"保存报告到 '{filename}'...")
        wb.save(filename)
        end_time = time.time()
        print(f"报告生成成功！耗时: {end_time - start_time:.2f} 秒。")

    except IOError as e:
        # More specific error for file saving issues
        print(f"错误：无法保存Excel报告到 '{filename}'。请检查文件是否已打开或权限问题。")
        print(f"  详细错误: {e}")
    except Exception as e:
        print(f"错误：生成Excel报告时发生意外错误: {str(e)}")
        # import traceback
        # traceback.print_exc() # Uncomment for detailed traceback during debugging


# --- Helper functions for creating specific report sheets ---

def _create_category_sheet(wb, videos):
    ws = wb.active
    ws.title = "分类统计"
    ws.append(["类别名称", "视频数量"])

    # 统计并清理类别名称
    category_counts = Counter()
    for v in videos:
        if v.category:
            clean_category = clean_excel_data(v.category)
            category_counts[clean_category] += 1

    # 按数量降序、名称升序排序
    for category, count in sorted(category_counts.items(), key=lambda item: (-item[1], item[0])):
        ws.append([category, count])

    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15


def _create_watch_history_sheet(wb, users):
    ws = wb.create_sheet("观看记录")

    # 生成日期表头
    base_date = datetime.now().date()
    date_headers = [(base_date - timedelta(days=i)).strftime("%m-%d")
                    for i in range(DAYS_TO_SIMULATE - 1, -1, -1)]
    ws.append(["用户ID"] + date_headers)

    # 处理每个用户的观看记录
    for user in users:
        daily_watches = defaultdict(list)
        for url, timestamp_str in user.watched_videos:
            try:
                day_key = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S').strftime("%m-%d")
                if day_key in date_headers:
                    clean_url = clean_excel_data(url)
                    daily_watches[day_key].append(clean_url)
            except (ValueError, KeyError):
                continue

        # 构建用户行数据
        user_row = [clean_excel_data(user.user_id)]
        for day_key in date_headers:
            urls_str = ", ".join(daily_watches.get(day_key, []))
            user_row.append(clean_excel_data(urls_str))
        ws.append(user_row)

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    for col in range(2, len(date_headers) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 25


def _create_like_history_sheet(wb, users):
    ws = wb.create_sheet("点赞记录")
    ws.append(["用户ID", "视频URL", "点赞时间"])

    for user in users:
        for url, timestamp_str in user.liked_videos:
            ws.append([
                clean_excel_data(user.user_id),
                clean_excel_data(url),
                clean_excel_data(timestamp_str)
            ])

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20


def _create_video_stats_sheet(wb, videos):
    ws = wb.create_sheet("视频统计")
    headers = ["视频URL", "标题", "封面地址", "播放地址", "类别", "观看次数", "点赞数"]
    ws.append(headers)

    # 按URL排序视频
    sorted_videos = sorted([v for v in videos if v.url != "N/A"], key=lambda v: v.url)

    for video in sorted_videos:
        ws.append([
            clean_excel_data(video.url),
            clean_excel_data(video.title),
            clean_excel_data(video.cover_url),
            clean_excel_data(video.play_url),
            clean_excel_data(video.category),
            len(video.viewers),
            video.like_count
        ])

    # 设置列宽
    column_widths = [50, 40, 40, 40, 30, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _create_daily_views_sheet(wb, videos, daily_views_data):
    ws = wb.create_sheet("每日观看")

    if not daily_views_data:
        ws.append(["视频URL", "无数据"])
        return

    # 获取排序后的日期
    sample_url = next(iter(daily_views_data), None)
    if not sample_url:
        ws.append(["视频URL", "无数据"])
        return

    dates = sorted(daily_views_data[sample_url].keys())
    ws.append(["视频URL"] + dates)

    # 按URL排序视频
    sorted_video_urls = sorted([v.url for v in videos if v.url != "N/A" and v.url in daily_views_data])

    for url in sorted_video_urls:
        row_data = [clean_excel_data(url)]
        row_data.extend([daily_views_data[url].get(date, 0) for date in dates])
        ws.append(row_data)

    # 设置列宽
    ws.column_dimensions['A'].width = 50
    for col in range(2, len(dates) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _create_hot_ranking_sheet(wb, videos, ranking_data):
    ws = wb.create_sheet("热门排行")

    if not ranking_data:
        # Adjust header slightly for consistency if no data
        headers = ["排名", "视频URL", "标题", "封面地址", "播放地址", "类别",
                   "观看数 (独立用户)", "点赞数", "热度评分", "无数据"] # Corrected header name
        ws.append(headers)
        return

    # 准备表头 - Use the correct key name here
    headers = ["排名", "视频URL", "标题", "封面地址", "播放地址", "类别",
               "观看数 (独立用户)", "点赞数", "热度评分"] # Corrected header name
    ws.append(headers)

    # 创建视频URL到视频对象的映射
    video_dict = {v.url: v for v in videos if v.url != "N/A"}

    for item in ranking_data:
        video = video_dict.get(item["视频URL"])
        row_data = [
            item["排名"],
            clean_excel_data(item["视频URL"]),
            clean_excel_data(video.title) if video else "",
            clean_excel_data(video.cover_url) if video else "",
            clean_excel_data(video.play_url) if video else "",
            clean_excel_data(item["类别"]),
            item["观看数 (独立用户)"], # <--- Use the correct key from ranking_data
            item["点赞数"], # This key should be correct
            item["热度评分"] # This key should be correct
        ]
        ws.append(row_data)

    # 设置列宽 - Adjust width for the potentially longer header
    # Indices: 1   2   3   4   5   6   7                    8    9
    column_widths = [8, 50, 40, 40, 40, 30, 20,               12,  12] # Increased width for "观看数 (独立用户)"
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


    # 设置列宽
    column_widths = [8, 50, 40, 40, 40, 30, 12, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


import pandas as pd


def save_videos_to_csv(videos, output_csv_file):
    """
    将视频数据保存到 CSV 文件。

    Args:
        videos (list): 包含 Video 对象的列表。
        output_csv_file (str): 输出 CSV 文件的路径。
    """
    # 先提取视频对象的相关数据
    video_data = []
    for video in videos:
        video_data.append({
            "视频ID": video.id,
            "视频URL": video.url,
            "关键词": " ".join(video.keywords),  # 将关键词转为字符串
            "类别": video.category,
            "标题": video.title,
            "点赞数": video.like_count,
            "封面URL": video.cover_url,
            "播放URL": video.play_url,
            "观看者数": len(video.viewers),  # 观看者数
            "点赞用户数": len(video.liked_users),  # 点赞用户数
        })

    # 创建 DataFrame
    df = pd.DataFrame(video_data)

    # 保存为 CSV 文件
    try:
        df.to_csv(output_csv_file, index=False, encoding='utf-8-sig')  # 使用 utf-8-sig 编码确保支持中文
        print(f"视频数据已成功保存到 {output_csv_file}")
    except Exception as e:
        print(f"保存视频数据到 CSV 时发生错误: {e}")


import pandas as pd


def save_users_to_csv(users, output_filename="users_output.csv"):
    """将用户信息保存到CSV文件"""

    # 生成用户信息的列表 (可以包含用户 ID 和兴趣等字段)
    users_data = []
    for user in users:
        user_info = {
            "用户ID": user.user_id,
            "兴趣类别": ", ".join(user.interests.keys()),  # 假设兴趣类别是一个字典
            "观看视频数": len(user.watched_videos),
            "点赞视频数": len(user.liked_videos)
        }
        users_data.append(user_info)

    # 创建 DataFrame
    df = pd.DataFrame(users_data)

    # 保存到 CSV 文件
    try:
        df.to_csv(output_filename, index=False, encoding='utf-8-sig')  # 使用 utf-8-sig 编码确保中文支持
        print(f"用户信息已成功保存到 '{output_filename}'")
    except Exception as e:
        print(f"保存用户信息到CSV时发生错误: {e}")


def load_clusters_from_excel_with_dtypes(excel_filepath, sheet_name, expected_dtypes=None):
    """
    从 Excel 文件加载指定工作表的数据到 Pandas DataFrame。
    如果提供了 expected_dtypes，则会尝试在加载时应用这些数据类型。

    参数:
        excel_filepath (str): Excel 文件的路径。
        sheet_name (str): 要读取的工作表的名称。
        expected_dtypes (dict, optional): 一个字典，键是列名，值是该列期望的 pandas/numpy 数据类型。
                                          例如: {'聚类ID': 'int64', '视频URL': 'object'}
                                          如果为 None，pandas 将自动推断数据类型。

    返回:
        pd.DataFrame: 加载后的 DataFrame。如果文件未找到或加载失败，则返回一个空的 DataFrame。
    """
    if not os.path.exists(excel_filepath):
        print(f"错误：Excel文件 '{excel_filepath}' 未找到。")
        return pd.DataFrame()

    try:
        print(f"正在从 Excel 文件 '{excel_filepath}' 的工作表 '{sheet_name}' 加载数据...")
        if expected_dtypes:
            print(f"  尝试使用预定义的数据类型: {expected_dtypes}")
            # 使用 dtype 参数来指定列的数据类型
            # 对于 .xlsx 文件，默认引擎通常是 'openpyxl'，确保已安装 (pip install openpyxl)
            reloaded_df = pd.read_excel(excel_filepath, sheet_name=sheet_name, dtype=expected_dtypes)
        else:
            print("  将由 pandas 自动推断数据类型。")
            reloaded_df = pd.read_excel(excel_filepath, sheet_name=sheet_name)

        print(f"成功从Excel加载 {len(reloaded_df)} 条记录。")
        print("加载后的DataFrame信息 (包含数据类型):")
        reloaded_df.info() # .info() 会显示每列的数据类型
        return reloaded_df
    except FileNotFoundError:
        print(f"错误：Excel文件 '{excel_filepath}' 未找到（在尝试读取时）。")
        return pd.DataFrame()
    except ValueError as ve: # 通常是 sheet_name 找不到
        print(f"从Excel文件 '{excel_filepath}' 加载数据时发生值错误: {ve}")
        print(f"  请确保工作表名称 '{sheet_name}' 正确无误。")
        return pd.DataFrame()
    except Exception as e:
        print(f"从Excel文件 '{excel_filepath}' 的工作表 '{sheet_name}' 加载数据时发生其他错误: {e}")
        return pd.DataFrame()
# ==================== 主程序 ====================
if __name__ == "__main__":
    main_start_time = time.time()
    print("=============================================")
    print("=== 视频用户行为模拟系统 - V2.1 (ID Based) ===") # Version bump
    print("=============================================")

    # --- 定义变量以存储中间结果 ---
    videos = []
    users = []
    video_clusters = pd.DataFrame() # Initialize as empty DataFrame
    user_clusters = pd.DataFrame()  # Initialize as empty DataFrame

    try:
        # === Step 0: Prepare Data Source ===
        print(f"\n--- [步骤 0] 数据源准备 ---")
        # ... (逻辑保持不变) ...
        print(f"源 Excel: {SOURCE_EXCEL_PATH}")
        print(f"目标 CSV: {TARGET_CSV_PATH}")

        csv_ready = False
        if os.path.exists(TARGET_CSV_PATH):
            print(f"找到已存在的CSV文件: '{os.path.basename(TARGET_CSV_PATH)}'")
            csv_ready = True
        else:
            print(f"未找到CSV文件，尝试从Excel转换...")
            if not os.path.exists(SOURCE_EXCEL_PATH):
                raise FileNotFoundError(f"错误：源Excel文件 '{SOURCE_EXCEL_PATH}' 不存在，无法进行转换。")

            if convert_excel_to_csv(SOURCE_EXCEL_PATH, TARGET_CSV_PATH):
                print(f"Excel成功转换为CSV。")
                csv_ready = True
            else:
                raise RuntimeError("Excel到CSV转换失败，无法继续。")

        if not csv_ready:
             print("错误：未能准备好CSV数据文件，程序终止。")
             sys.exit(1)

        # === Step 1: Load Data from CSV ===
        print(f"\n--- [步骤 1] 加载视频数据 ---")
        videos = load_csv_data(TARGET_CSV_PATH) # 确保加载函数处理ID
        if not videos:
             print("错误：未能从CSV加载任何视频数据，程序终止。")
             sys.exit(1)
        # print(f"成功加载 {len(videos)} 个视频。") # Message now printed inside load_csv_data

        # === Step 2: Categorize Videos ===
        print(f"\n--- [步骤 2] 视频分类 ---")
        # ... (分类恢复和自动分类逻辑保持不变) ...
        print("尝试从之前的报告文件中恢复分类信息...")
        report_files_to_check = [
            OUTPUT_FILENAME, # Main report file
            "video_clusters.xlsx", # Video clustering report might have categories
            "video_clusters_optimized_balanced.xlsx" # Add the new optimized filename
        ]
        try:
             categories_recovered, videos_updated_count = category_recovery.recover_categories(videos, report_files_to_check)
             print(f"分类恢复尝试完成。更新了 {videos_updated_count} 个视频的分类。")
        except NameError:
             print("警告：`category_recovery` 模块或功能未找到，跳过分类恢复。")
             categories_recovered = None; videos_updated_count = 0
        except Exception as rec_err:
             print(f"警告：在恢复分类时发生错误: {rec_err}。跳过恢复。")
             categories_recovered = None; videos_updated_count = 0

        categories_assigned_count = sum(1 for v in videos if hasattr(v, 'category') and v.category is not None and v.category not in ["N/A", ""])
        needs_auto_categorize = True
        # Adjust threshold or logic if needed
        if categories_assigned_count >= len(videos) * 0.7:
            print(f"已成功恢复或分配分类给 {categories_assigned_count}/{len(videos)} 个视频。跳过自动分类。")
            needs_auto_categorize = False
        else:
            print(f"当前分类信息不足 ({categories_assigned_count}/{len(videos)} 个视频)。")
            print("执行自动分类...")

        category_counts_result = {}
        if needs_auto_categorize:
            category_counts_result = auto_categorize(videos, n_categories=400) # Ensure n_categories is passed if needed by the func

        # 准备用于用户兴趣生成的有效类别列表
        all_assigned_categories = set(v.category for v in videos if hasattr(v, 'category') and v.category)
        valid_categories_for_interests = [
            cat for cat in all_assigned_categories
            if cat and cat not in ["No Keywords", "Other", "Clustering Error", "Single Cluster", "N/A", None] and str(cat).startswith("CAT_")
        ]
        if not valid_categories_for_interests and category_counts_result:
            valid_categories_for_interests = [cat for cat in category_counts_result if str(cat).startswith("CAT_")]

        if not valid_categories_for_interests:
            print("警告：分类后未找到有效的 'CAT_XXX' 簇类别。用户兴趣将基于 '默认类别'。")
            # User class handles the default, no need to set it here unless overriding

        print(f"用于用户兴趣生成的有效类别数量: {len(valid_categories_for_interests)}")


        # === Step 3: Generate Users ===
        print(f"\n--- [步骤 3] 生成模拟用户 ---")
        print(f"创建 {NUM_USERS} 个模拟用户...")
        users = [User(i, valid_categories_for_interests) for i in range(NUM_USERS)]
        print(f"成功创建 {len(users)} 个用户。")


        # === Step 4: Simulate User Behavior ===
        print(f"\n--- [步骤 4] 模拟用户行为 ---")
        simulate_behavior(users, videos)


        # === Step 5: Data Validation ===
        print(f"\n--- [步骤 5] 数据验证 ---")
        # ... (逻辑保持不变) ...
        total_watches_recorded = sum(len(u.watched_videos) for u in users)
        total_likes_recorded = sum(len(u.liked_videos) for u in users)
        videos_with_zero_views = [v.url for v in videos if hasattr(v,'viewers') and not v.viewers and getattr(v, 'url', 'N/A') != "N/A"]
        users_with_zero_watches = [u.user_id for u in users if not u.watched_videos]

        print(f"总记录观看次数: {total_watches_recorded}")
        print(f"总记录点赞次数: {total_likes_recorded}")
        if total_watches_recorded > 0:
             print(f"整体点赞率: {total_likes_recorded / total_watches_recorded:.2%}")
        else:
             print("整体点赞率: N/A (无观看记录)")
        print(f"模拟后仍无观看记录的视频数: {len(videos_with_zero_views)}")
        if videos_with_zero_views:
             print(f"  (注意: 这可能包括未被选中强制观看的视频，如果存在)")
        print(f"模拟后无观看记录的用户数: {len(users_with_zero_watches)}")


        # === Step 6: Generate Detailed Report ===
        print(f"\n--- [步骤 6] 生成详细报告 ---")
        save_detailed_report(users, videos, OUTPUT_FILENAME)


        # === Step 7: Advanced Clustering Analysis ===
        print(f"\n--- [步骤 7] 高级聚类分析 ---")

        # F6: 视频聚类分析
        try:
            video_cluster_start_time = time.time()
            print("\n执行视频聚类分析 (基于用户观看行为)...")
            # --- 使用最新的聚类函数 ---
            # Ensure the clustering function returns a DataFrame or None
            temp_video_clusters = video_clustering.cluster_videos_by_viewers_balanced_optimized(
                videos=videos,
                users=users,
                n_clusters=100, # Adjust as needed
                batch_size=5000,
                max_size_factor=1.8,
                split_k=2,
                svd_components=150, # Or 0 to disable SVD
                mbk_n_init=5
            )
            # --- 存储结果 ---
            if isinstance(temp_video_clusters, pd.DataFrame):
                 video_clusters = temp_video_clusters # Store the result
                 print(f"视频聚类完成，生成了 {len(video_clusters)} 条记录的 DataFrame。")
            else:
                 print("视频聚类未返回有效的 DataFrame。")
                 video_clusters = pd.DataFrame() # Keep it as an empty DataFrame

            video_cluster_end_time = time.time()
            print(f"  >>> 视频聚类耗时: {video_cluster_end_time - video_cluster_start_time:.2f} 秒 <<<")

            # 保存结果 (仅当 DataFrame 非空时)
            if not video_clusters.empty:
                video_cluster_file = "video_clusters_optimized_balanced.xlsx"
                print(f"  准备将聚类结果保存到 '{video_cluster_file}'...")
                try:
                    with pd.ExcelWriter(video_cluster_file) as writer:
                        # 保存详情
                        video_clusters.to_excel(writer, sheet_name="视频聚类详情", index=False)
                        print(f"    已写入 '视频聚类详情' sheet ({len(video_clusters)} 行)。")

                        # 保存统计 (确保列存在)
                        print("    计算并写入 '聚类统计' sheet...")
                        required_stat_cols = ["聚类ID", "视频URL", "观看该视频的用户数", "聚类内视频数"]
                        if all(col in video_clusters.columns for col in required_stat_cols):
                            cluster_stats = video_clusters.groupby("聚类ID").agg(
                                视频数量=("视频URL", "count"),
                                平均观看用户数_每个视频=("观看该视频的用户数", "mean"),
                                实际簇大小=("聚类内视频数", "first") # Assuming this column exists and is consistent per cluster
                            ).reset_index()
                            cluster_stats.to_excel(writer, sheet_name="聚类统计", index=False)
                            print(f"      写入了 {len(cluster_stats)} 个聚类的统计信息。")
                        else:
                             missing_stat_cols = [c for c in required_stat_cols if c not in video_clusters.columns]
                             print(f"    警告：无法生成聚类统计，结果 DataFrame 缺少列: {missing_stat_cols}。")

                    print(f"  视频聚类结果已保存到 {video_cluster_file}")
                except Exception as save_err:
                    print(f"  错误：保存视频聚类结果到Excel时失败: {save_err}")
                    # Optional: Save as CSV backup
                    try:
                        csv_path = os.path.splitext(video_cluster_file)[0] + ".csv"
                        video_clusters.to_csv(csv_path, index=False, encoding='utf-8-sig')
                        print(f"  已尝试将结果保存为 CSV: {csv_path}")
                    except Exception as csv_save_err:
                        print(f"  错误：尝试保存为 CSV 也失败了: {csv_save_err}")
            else:
                 print("  视频聚类结果为空，跳过保存。")

        except AttributeError as ae:
            print(f"\n视频聚类失败：属性错误 - {ae}")
            print("  请确保 'video_clustering.py' 文件中存在所需的函数和属性。")
            video_clusters = pd.DataFrame() # Ensure it's empty on error
        except Exception as e:
            print(f"\n视频聚类或结果保存时发生错误: {str(e)}")
            print("详细错误追踪:")
            traceback.print_exc()
            video_clusters = pd.DataFrame() # Ensure it's empty on error


        # F7: 用户聚类分析
        try:
            print("\n执行用户聚类分析...")
            # --- 存储结果 ---
            temp_user_clusters = user_clustering.cluster_users_by_interests(users, videos, n_clusters=100)
            if isinstance(temp_user_clusters, pd.DataFrame):
                 user_clusters = temp_user_clusters
                 print(f"用户聚类完成，生成了 {len(user_clusters)} 条记录的 DataFrame。")
            else:
                 print("用户聚类未返回有效的 DataFrame。")
                 user_clusters = pd.DataFrame()

            # 保存结果 (仅当 DataFrame 非空时)
            if not user_clusters.empty:
                user_cluster_file = "user_clusters.xlsx"
                with pd.ExcelWriter(user_cluster_file) as writer:
                    user_clusters.to_excel(writer, sheet_name="用户聚类", index=False)
                    # 添加聚类统计信息 (确保列存在)
                    required_user_stat_cols = ["聚类ID", "用户ID", "观看类别数"]
                    if all(col in user_clusters.columns for col in required_user_stat_cols):
                         cluster_stats = user_clusters.groupby("聚类ID").agg(
                              用户数量=("用户ID", "count"),
                              平均类别数=("观看类别数", "mean")
                         ).reset_index() # Add reset_index for cleaner output
                         cluster_stats.to_excel(writer, sheet_name="聚类统计", index=False) # Save stats without index
                    else:
                         missing_user_stat_cols = [c for c in required_user_stat_cols if c not in user_clusters.columns]
                         print(f"    警告：无法生成用户聚类统计，结果 DataFrame 缺少列: {missing_user_stat_cols}。")

                print(f"用户聚类结果已保存到 {user_cluster_file}")
            else:
                 print("  用户聚类结果为空，跳过保存。")

        except AttributeError as ae:
             print(f"用户聚类失败：属性错误 - {ae}")
             print("  请确保 'user_clustering.py' 文件中存在所需的函数和属性。")
             user_clusters = pd.DataFrame() # Ensure it's empty on error
        except Exception as e:
            print(f"用户聚类或保存失败: {str(e)}")
            traceback.print_exc()
            user_clusters = pd.DataFrame() # Ensure it's empty on error

        # === Step 8: User Center API Function Examples ===
        # ... (用户中心函数调用保持不变) ...
        print("\n--- [步骤 8] 用户中心 API 函数示例调用 ---")
        if users and videos:
            target_user_idx = 5
            if 0 <= target_user_idx < len(users):
                 print(f"\n获取用户索引 {target_user_idx} ({users[target_user_idx].user_id}) 的观看历史:")
                 watched_history_json = user_center.get_user_history(target_user_idx, users, videos, history_type='watched')
                 print(watched_history_json)
            else: print(f"用户索引 {target_user_idx} 无效。")

            target_user_idx = 10
            if 0 <= target_user_idx < len(users):
                 print(f"\n获取用户索引 {target_user_idx} ({users[target_user_idx].user_id}) 的点赞历史:")
                 liked_history_json = user_center.get_user_history(target_user_idx, users, videos, history_type='liked')
                 print(liked_history_json)
            else: print(f"用户索引 {target_user_idx} 无效。")

            if not user_clusters.empty: # Check if user clustering was successful
                target_user_idx = 5
                if 0 <= target_user_idx < len(users):
                     print(f"\n获取与用户索引 {target_user_idx} ({users[target_user_idx].user_id}) 相似的用户 (Top 5):")
                     similar_users_json = user_center.get_similar_users(target_user_idx, users, user_clusters, top_n=5)
                     print(similar_users_json)
                else: print(f"用户索引 {target_user_idx} 无效。")
            else:
                print("\n无法获取相似用户，因为用户聚类数据不可用或为空。")
        else:
            print("\n无法执行用户中心示例调用，因为用户或视频数据未加载。")


        # === [修改] Step 9: Video Playing API 函数调用测试 (使用 ID) ===
        print("\n--- [步骤 9] 视频播放/推荐 API 函数调用测试 (基于 ID) ---")

        # --- 前提检查 ---
        if 'videos' in locals() and videos:
            # --- 获取一些有效的视频 ID 用于测试 ---
            valid_ids_to_test = [v.id for v in videos[:5] if hasattr(v, 'id') and v.id] # Get first 5 valid IDs
            if not valid_ids_to_test:
                 # Try getting ID from later in the list if first few are bad
                 valid_ids_to_test = [v.id for v in videos[len(videos)//2 : len(videos)//2 + 5] if hasattr(v, 'id') and v.id]

            if valid_ids_to_test:
                test_play_id = valid_ids_to_test[0] # Use the first valid ID found
                print(f"\n[测试] 获取视频 ID '{test_play_id}' 的播放信息:")
                play_info_json_id = videos_playing.play_video_by_id(test_play_id, videos)
                print(play_info_json_id)

                # 测试一个理论上不存在的 ID
                test_play_invalid_id = "ID_绝对_不_存在_999"
                print(f"\n[测试] 获取视频 ID '{test_play_invalid_id}' 的播放信息 (无效 ID):")
                play_info_json_invalid_id = videos_playing.play_video_by_id(test_play_invalid_id, videos)
                print(play_info_json_invalid_id)

                # --- 测试 get_similar_videos_by_id ---
                # 检查 video_clusters 是否有效
                if 'video_clusters' in locals() and isinstance(video_clusters, pd.DataFrame) and not video_clusters.empty:
                    test_similar_id = valid_ids_to_test[-1] # Use the last valid ID found
                    num_to_find = 5
                    print(f"\n[测试] 获取与视频 ID '{test_similar_id}' 相似的 {num_to_find} 个视频:")
                    similar_videos_json_id = videos_playing.get_similar_videos_by_id(
                        test_similar_id,
                        videos,
                        video_clusters, # 传递聚类结果 DataFrame
                        num_similar=num_to_find
                    )
                    print(similar_videos_json_id)

                    # 测试一个理论上不存在的 ID
                    test_similar_nonexistent_id = "ID_绝对_不_存在_用于相似度_888"
                    print(f"\n[测试] 获取与视频 ID '{test_similar_nonexistent_id}' 相似的视频 (无效 ID):")
                    similar_videos_json_invalid_id = videos_playing.get_similar_videos_by_id(
                        test_similar_nonexistent_id,
                        videos,
                        video_clusters,
                        num_similar=num_to_find
                    )
                    print(similar_videos_json_invalid_id)
                else:
                    print("\n错误：无法执行 get_similar_videos_by_id 测试，因为 'video_clusters' DataFrame 不存在、无效或为空。")
                    print(f"  (video_clusters 类型: {type(video_clusters)}, 是否为空: {video_clusters.empty if isinstance(video_clusters, pd.DataFrame) else 'N/A'})")

            else:
                print("\n错误：在 'videos' 列表中找不到有效的视频 ID 用于 API 函数测试。")
        else:
            print("\n错误：'videos' 列表不存在或为空，无法执行视频播放/推荐 API 函数测试。")

        # 使用示例
        output_csv_file = "videos_output.csv"
        save_videos_to_csv(videos, output_csv_file)
        # 示例：调用该函数并传入 `users` 列表
        save_users_to_csv(users)
        # 这是您的主脚本保存的Excel文件的路径
        cluster_excel_file_path = "video_clusters_optimized_balanced.xlsx"
        # 这是您要读取的工作表的名称
        target_sheet_name = "视频聚类详情"  # 主数据通常在这个sheet

        expected_column_dtypes = {
            "视频URL": "object",  # 'object' 通常用于字符串
            "聚类ID": "int64",  # 或者 'int32' 如果数值范围允许
            "观看该视频的用户数": "int64",  # 或者 'int32'
            "聚类内视频数": "int64"  # 或者 'int32'
        }
        # 确保这里的列名与您的Excel文件 "视频聚类详情" 工作表中的列名完全匹配。

        print(f"\n--- 场景 1: 从Excel '{target_sheet_name}' 工作表使用期望的数据类型加载 ---")

        # 为了使这个示例能独立运行，如果Excel文件不存在，我们创建一个虚拟的
        if not os.path.exists(cluster_excel_file_path):
            print(f"警告: '{cluster_excel_file_path}' 不存在。将创建一个虚拟Excel文件用于测试。")
            # 确保虚拟数据的列与 expected_column_dtypes 中的键匹配
            dummy_data_details = {
                "视频URL": ["http://example.com/vid1_excel", "http://example.com/vid2_excel",
                            "http://example.com/vid3_excel"],
                "聚类ID": [10, 11, 10],
                "观看该视频的用户数": [1000, 500, 1200],
                "聚类内视频数": [20, 10, 20]
            }
            dummy_df_details = pd.DataFrame(dummy_data_details)

            # 您的Excel文件有两个sheet，我们也创建两个
            dummy_data_stats = {  # 假设的统计数据表结构 (与主脚本中的一致)
                "聚类ID": [10, 11],
                "视频数量": [2, 1],
                "平均观看用户数_每个视频": [1100.0, 500.0],  # 注意这个可能是 float
                "实际簇大小": [20, 10]
            }
            dummy_df_stats = pd.DataFrame(dummy_data_stats)

            try:
                # 需要安装 openpyxl: pip install openpyxl
                with pd.ExcelWriter(cluster_excel_file_path, engine='openpyxl') as writer:
                    dummy_df_details.to_excel(writer, sheet_name="视频聚类详情", index=False)
                    dummy_df_stats.to_excel(writer, sheet_name="聚类统计", index=False)
                print(f"虚拟Excel文件 '{cluster_excel_file_path}' 已创建。")
            except Exception as ex_create_err:
                print(f"创建虚拟Excel文件失败: {ex_create_err}")
                print("  请确保已安装 'openpyxl' 库 (pip install openpyxl)。")

        reloaded_clusters_from_excel_with_types = load_clusters_from_excel_with_dtypes(
            excel_filepath=cluster_excel_file_path,
            sheet_name=target_sheet_name,
            expected_dtypes=expected_column_dtypes
        )

        if not reloaded_clusters_from_excel_with_types.empty:
            print(f"\n从Excel '{target_sheet_name}' 重新加载的视频聚类数据 (指定了Dtypes):")
            print(reloaded_clusters_from_excel_with_types.head())
            print("\n数据类型:")
            print(reloaded_clusters_from_excel_with_types.dtypes)  # 再次确认数据类型
    # === 异常处理和结束 ===
    except FileNotFoundError as fnf:
        print(f"致命错误：找不到所需文件! {fnf}", file=sys.stderr)
        print("请确认所有输入文件路径正确，或检查相关目录权限。")
    except PermissionError as pe:
        print(f"致命错误：无权限访问所需文件或文件夹! {pe}", file=sys.stderr)
        print("请确认您有足够的系统权限，或者相关文件未被其他程序占用。")
    except MemoryError:
        print("致命错误：系统内存不足!", file=sys.stderr)
        print("提示: 尝试减少处理的数据量，关闭其他应用程序，或在更多内存的计算机上运行。")
    except KeyboardInterrupt:
        print("\n程序被用户中断!", file=sys.stderr)
        print("部分处理结果可能已保存。")
    except RuntimeError as rte: # Catch specific runtime errors raised
         print(f"运行时错误: {rte}", file=sys.stderr)
    except Exception as e:
        print(f"发生未预期的致命错误: {e.__class__.__name__} - {e}", file=sys.stderr)
        print("详细错误追踪:", file=sys.stderr)
        traceback.print_exc(file=sys.stderr) # Print full traceback for unexpected errors
    finally:
        main_end_time = time.time()
        print(f"\n程序总执行时间: {main_end_time - main_start_time:.2f} 秒。")
        print("程序退出。")
        # Any cleanup code here if needed
        gc.collect() # Final garbage collection attempt
