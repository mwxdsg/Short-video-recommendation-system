# -*- coding: utf-8 -*-
"""
视频用户行为模拟系统

功能概述：
1. 从Excel加载视频数据（包含URL和关键词）
2. 使用TF-IDF和余弦相似度对视频进行自动分类
3. 生成模拟用户及其兴趣偏好
4. 模拟用户的观看和点赞行为
5. 生成详细的行为报告和统计数据
"""

# ==================== 导入依赖库 ====================
import random
import os
from datetime import datetime
from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook  # Excel文件处理
from sklearn.feature_extraction.text import TfidfVectorizer  # TF-IDF计算
from sklearn.metrics.pairwise import cosine_similarity  # 余弦相似度计算
import numpy as np
from tqdm import tqdm  # 进度条显示

# ==================== 全局配置参数 ====================
NUM_USERS = 1000  # 模拟用户数量
MIN_INTERESTS = 4  # 每个用户最少兴趣类别数
MAX_INTERESTS = 6  # 每个用户最多兴趣类别数
OUTPUT_FILENAME = "user_behavior_details_tfidf.xlsx"  # 输出文件名
SIMILARITY_THRESHOLD = 0.5  # 关键词聚类相似度阈值


# ==================== 核心类定义 ====================
class Video:
    """视频对象，包含元数据和行为统计"""

    def __init__(self, url, keywords):
        self.url = url.strip()  # 视频唯一标识
        self.keywords = [kw for kw in (keywords.split() if keywords else []) if kw]  # 空格分割并过滤空值
        self.category = None  # 自动分类后的类别（通过TF-IDF计算得出）
        self.like_count = 0  # 累计获得的点赞数
        self.viewers = set()  # 观看过该视频的用户ID集合
        self.liked_users = set()  # 新增点赞用户记录

class User:
    """用户对象，包含兴趣偏好和行为记录"""

    def __init__(self, user_id, all_categories):
        self.user_id = f"U{user_id:04d}"  # 用户ID格式化为U0001样式
        self.interests = self._generate_interests(all_categories)  # 用户兴趣字典：{类别: 兴趣权重}
        self.watched_videos = []  # 观看记录列表，元素为元组（视频URL, 时间戳）
        self.liked_videos = []  # 点赞记录列表，元素为元组（视频URL, 时间戳）

    def _generate_interests(self, categories):
        """随机生成用户兴趣分布"""
        if not categories:
            return {}
        # 随机选择兴趣类别数量（在MIN和MAX之间）
        num_interests = random.randint(
            min(MIN_INTERESTS, len(categories)),
            min(MAX_INTERESTS, len(categories))
        )
        chosen_cats = random.sample(categories, num_interests)
        # 为每个兴趣类别生成随机权重（0.5-1.0之间）
        return {cat: random.uniform(0.7, 1.0) for cat in chosen_cats}


# ==================== 数据处理函数 ====================
def load_excel_data(file_path):
    """从Excel加载视频数据
    Args:
        file_path: 输入文件路径，要求包含url和keywords列
    Returns:
        Video对象列表
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Input file {file_path} not found")

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        videos = []

        # 自动检测列位置（不区分大小写）
        col_mapping = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and 'url' in header.lower():
                col_mapping['url'] = col
            elif header and ('keyword' in header.lower() or 'tag' in header.lower()):
                col_mapping['keywords'] = col

        if not col_mapping:
            raise ValueError("Excel文件中未找到必要的列（url/keywords）")

        # 逐行读取视频数据
        for row in range(2, ws.max_row + 1):
            url = ws.cell(row=row, column=col_mapping['url']).value
            keywords = ws.cell(row=row, column=col_mapping['keywords']).value or ""
            if url:
                videos.append(Video(url, keywords))

        return videos
    except Exception as e:
        raise ValueError(f"Excel文件读取错误: {str(e)}")


# 修改后的auto_categorize函数（替换原函数）
def auto_categorize(videos):
    """使用层次聚类强制最多400个分类"""
    if not videos:
        print("警告: 没有可供分类的视频数据")
        return {}

    # 生成视频特征文档（合并每个视频的关键词）
    print("正在生成视频特征文档...")
    video_docs = [' '.join(v.keywords) for v in videos if v.keywords]

    if not video_docs:
        print("警告: 所有视频均无有效关键词")
        return {}

    # 计算TF-IDF矩阵（优化参数提升效率）
    print("正在计算TF-IDF特征(1-2 grams)...")
    tfidf = TfidfVectorizer(
        ngram_range=(1, 2),
        min_df=3,  # 提高最小出现次数阈值
        max_features=5000,  # 限制最大特征数量
        token_pattern=r'\b\w+\b'
    )
    try:
        tfidf_matrix = tfidf.fit_transform(video_docs)
    except ValueError:
        print("警告: 特征维度不足，无法进行分类")
        return {}

    # 优化聚类配置
    MAX_CATEGORIES = 400  # 强制最多200个分类
    from sklearn.cluster import AgglomerativeClustering

    # 动态调整簇数量（不超过视频总量）
    actual_clusters = min(MAX_CATEGORIES, len(videos))

    print(f"正在进行层次聚类（强制最多{actual_clusters}类）...")
    clustering = AgglomerativeClustering(
        n_clusters=actual_clusters,  # 强制指定簇数量
        metric='cosine',  # 直接使用余弦相似度
        linkage='average'
    )
    clusters = clustering.fit_predict(tfidf_matrix.toarray())

    # 生成类别标签（优化标签生成规则）
    print("生成类别标签...")
    category_map = {}
    cluster_keywords = defaultdict(list)

    # 收集每个簇的关键词（按TF-IDF权重）
    feature_names = tfidf.get_feature_names_out()
    for idx, cluster_id in enumerate(clusters):
        # 获取当前视频的权重向量
        vec = tfidf_matrix[idx]
        # 提取权重非零的特征
        feature_weights = zip(vec.indices, vec.data)
        # 按权重排序并保留前5个关键词
        sorted_features = sorted(feature_weights, key=lambda x: -x[1])[:3]
        selected_keywords = [feature_names[i] for i, _ in sorted_features]
        cluster_keywords[cluster_id].extend(selected_keywords)

    # 创建可读性高的类别名称
    for cluster_id, keywords in cluster_keywords.items():
        # 统计词频（考虑权重因素）
        counter = Counter(keywords)
        top_keywords = [kw for kw, _ in counter.most_common(3)]
        # 生成短标签（最多两个关键词）
        label = "_".join(top_keywords[:2]) if top_keywords else "其他"
        category_map[cluster_id] = f"CAT_{label}"

    # 分配视频类别
    print("分配视频类别...")
    category_counter = Counter()
    for idx, video in enumerate(videos):
        if video.keywords:
            cluster_id = clusters[idx]
            video.category = category_map.get(cluster_id, "其他")
            category_counter[video.category] += 1
        else:
            video.category = "无关键词"
            category_counter[video.category] += 1

    # 打印分类统计
    print("\n=== 分类结果统计（最多400类） ===")
    print("{:<25} {:<10}".format("类别名称", "视频数量"))
    print("-" * 45)
    for cat, count in category_counter.most_common(15):  # 仅显示前15大类
        print("{:<25} {:<10}".format(cat, count))
    print(f"（总共 {len(category_counter)} 个分类）")

    return category_map
# ==================== 统计与展示函数 ====================
def list_all_categories(videos):
    """展示视频分类统计信息"""
    category_stats = defaultdict(int)
    for video in videos:
        if video.category:
            category_stats[video.category] += 1

    print("\n=== 视频类别统计 ===")
    print("{:<30} {:<10}".format("类别名称", "视频数量"))
    print("-" * 45)
    for category, count in sorted(category_stats.items(),
                                  key=lambda x: x[1], reverse=True):
        print("{:<30} {:<10}".format(category, count))

    return category_stats


# ==================== 用户行为模拟函数 ====================
def simulate_behavior(users, videos):
    """模拟用户观看和点赞行为
    Logic:
        1. 根据用户兴趣权重选择视频类别
        2. 每个用户进行多个会话（session）
        3. 每个会话中观看多个视频
        4. 基于兴趣权重计算点赞概率
    """
    if not users or not videos:
        print("警告: 缺少用户或视频数据")
        return

    # 按类别建立视频索引
    cat_index = defaultdict(list)
    for v in videos:
        if v.category:
            cat_index[v.category].append(v)

    # 遍历所有用户模拟行为
    for user in tqdm(users, desc="用户行为模拟中"):
        # 随机生成1-3个会话
        for _ in range(random.randint(2, 4)):
            # 生成会话时间（当天8-23点间随机）
            session_time = datetime.now().replace(
                hour=random.randint(8, 23),
                minute=random.randint(0, 59))

            # 每个会话观看70-100个视频
            for _ in range(random.randint(70, 100)):
                # 90%概率按兴趣选择视频，10%随机选择
                if user.interests and random.random() < 0.9:
                    # 按兴趣权重选择类别
                    chosen_cat = random.choices(
                        list(user.interests.keys()),
                        weights=list(user.interests.values()),
                        k=1
                    )[0]
                    candidates = cat_index.get(chosen_cat, [])
                else:
                    candidates = videos

                if candidates:
                    video = random.choice(candidates)
                    _record_watch(user, video, session_time)

                    # 计算点赞概率（基础20% + 兴趣权重影响）
                    like_prob = 0.2
                    if video.category in user.interests:
                        like_prob += 0.8 * user.interests[video.category]

                    if random.random() < like_prob:
                        _record_like(user, video, session_time)


def _record_watch(user, video, time):
    """记录观看行为"""
    entry = (video.url, time.strftime('%Y-%m-%d %H:%M:%S'))
    user.watched_videos.append(entry)
    video.viewers.add(user.user_id)


def _record_like(user, video, time):
    """记录点赞行为"""
    if user.user_id not in video.liked_users:
        entry = (video.url, time.strftime('%Y-%m-%d %H:%M:%S'))
        user.liked_videos.append(entry)
        video.like_count += 1
        video.liked_users.add(user.user_id)


# ==================== 报告生成函数 ====================
def save_detailed_report(users, videos, filename):
    """生成Excel格式的详细报告
    包含4个工作表:
        1. 类别统计
        2. 观看记录
        3. 点赞记录
        4. 视频统计数据
    """
    try:
        wb = Workbook()

        # Sheet 1: 类别统计
        ws_cats = wb.active
        ws_cats.title = "Category Stats"
        ws_cats.append(["类别名称", "视频数量"])
        cat_stats = defaultdict(int)
        for v in videos:
            if v.category:
                cat_stats[v.category] += 1
        for cat, count in sorted(cat_stats.items(), key=lambda x: x[1], reverse=True):
            ws_cats.append([cat, count])

        # Sheet 2: 观看记录
        ws_watches = wb.create_sheet("Watch History")
        ws_watches.append(["用户ID", "视频URL", "观看时间"])
        for user in users:
            for url, time in user.watched_videos:
                ws_watches.append([user.user_id, url, time])

        # Sheet 3: 点赞记录
        ws_likes = wb.create_sheet("Like History")
        ws_likes.append(["用户ID", "视频URL", "点赞时间"])
        for user in users:
            for url, time in user.liked_videos:
                ws_likes.append([user.user_id, url, time])

        # Sheet 4: 视频统计数据
        ws_videos = wb.create_sheet("Video Statistics")
        ws_videos.append(["视频URL", "所属类别", "观看次数", "点赞次数"])
        for video in videos:
            view_count = sum(1 for u in users if any(url == video.url for url, _ in u.watched_videos))
            ws_videos.append([video.url, video.category, view_count, video.like_count])

        wb.save(filename)
    except Exception as e:
        raise IOError(f"Excel文件保存失败: {str(e)}")


# ==================== 主程序 ====================
if __name__ == "__main__":
    try:
        # 步骤1: 加载视频数据
        print("正在加载视频数据...")
        videos = load_excel_data("E:\数据结构大作业\数据1.xlsx")
        print(f"成功加载 {len(videos)} 条视频数据")

        # 步骤2: 自动分类视频
        print("正在使用TF-IDF进行视频分类...")
        cat_names = auto_categorize(videos)

        # 步骤3: 显示分类统计
        print("视频分类结果:")
        list_all_categories(videos)

        # 步骤4: 生成模拟用户
        print(f"正在创建 {NUM_USERS} 个模拟用户...")
        # 获取有效的分类标签（排除系统分类）
        valid_categories = [
            cat for cat in cat_names.values()
            if cat not in ["其他", "无关键词"]
        ]
        users = [User(i, valid_categories) for i in range(NUM_USERS)]
        # 步骤5: 模拟用户行为
        print("开始模拟用户行为...")
        simulate_behavior(users, videos)

        # 步骤6: 保存报告
        print(f"正在生成报告文件 {OUTPUT_FILENAME}...")
        save_detailed_report(users, videos, OUTPUT_FILENAME)

        # 汇总统计
        total_views = sum(len(u.watched_videos) for u in users)
        total_likes = sum(len(u.liked_videos) for u in users)

        print("\n=== 模拟结果摘要 ===")
        print(f"总观看次数: {total_views}")
        print(f"总点赞次数: {total_likes}")
        print(f"点赞率: {total_likes / total_views:.1%}")
        print(f"报告已保存至 {OUTPUT_FILENAME}")

    except FileNotFoundError as e:
        print(f"文件错误: {str(e)}")
    except ValueError as e:
        print(f"数据错误: {str(e)}")
    except Exception as e:
        print(f"未预期错误: {str(e)}")







# -*- coding: utf-8 -*-
""" 优化版视频用户行为模拟系统 """
import pandas as pd
from sklearn.cluster import MiniBatchKMeans
from joblib import Memory
import warnings
warnings.filterwarnings('ignore')

# ==================== 全局配置 ====================
OUTPUT_FILE = "optimized_behavior.xlsx"
CACHE_DIR = "./cache"

# ==================== 数据缓存 ====================
memory = Memory(CACHE_DIR, verbose=0)

@memory.cache
def load_data(file_path):
    """ 优化版数据加载 """
    # 只读取必要列并指定数据类型
    dtype_spec = {
        '用户ID': 'category', '视频URL': 'string', 
        '所属类别': 'category', '观看次数': 'uint32', 
        '点赞次数': 'uint16'
    }
    
    with pd.ExcelFile(file_path) as xls:
        watch = pd.read_excel(
            xls, sheet_name="Watch History",
            usecols=["用户ID", "视频URL", "观看时间"],
            dtype=dtype_spec
        )
        videos = pd.read_excel(
            xls, sheet_name="Video Statistics",
            usecols=["视频URL", "所属类别", "观看次数", "点赞次数"],
            dtype=dtype_spec
        )
    return watch, videos

# ==================== 核心优化 ====================
class Recommender:
    def __init__(self, watch_df, videos_df):
        # 建立索引
        self.watch = watch_df.set_index("用户ID")
        self.videos = videos_df.set_index("视频URL")
        
        # 预计算全局数据
        self.category_stats = videos_df["所属类别"].value_counts(normalize=True)
        self.max_views = videos_df["观看次数"].max()
        self.max_likes = videos_df["点赞次数"].max()
        
        # 用户相似度矩阵
        self.user_sim_matrix = self._precompute_user_similarity()
    
    def _precompute_user_similarity(self):
        """ 预计算用户相似度 """
        user_cat_matrix = (
            self.watch.join(self.videos, on="视频URL")
            .groupby(["用户ID", "所属类别"])
            .size()
            .unstack(fill_value=0)
        )
        return cosine_similarity(user_cat_matrix)

    def get_user_recommendations(self, user_id, top_n=10, diversify=True):
        """ 综合推荐引擎 """
        # 获取用户历史
        try:
            user_watched = self.watch.loc[user_id]["视频URL"].tolist()
            user_categories = self.videos.loc[user_watched]["所属类别"]
            cat_counts = user_categories.value_counts()
        except KeyError:
            return self._get_fallback_recommendations(top_n)
        
        # 候选视频
        candidates = self.videos[~self.videos.index.isin(user_watched)].copy()
        
        # 向量化评分
        candidates["base_score"] = (
            0.4 * candidates["所属类别"].map(cat_counts).fillna(0) +
            0.4 * candidates["观看次数"] / self.max_views +
            0.2 * candidates["点赞次数"] / self.max_likes
        )
        
        # 多样性增强
        if diversify:
            candidates = self._apply_diversity(candidates, user_categories)
        
        return candidates.nlargest(top_n, "final_score")

    def _apply_diversity(self, candidates, user_categories):
        """ 多样性增强策略 """
        # 1. 类别覆盖率奖励
        candidates["cat_coverage"] = ~candidates["所属类别"].isin(user_categories)
        # 2. 长尾内容奖励
        candidates["longtail"] = candidates["观看次数"] < self.max_views * 0.1
        # 3. 综合得分
        candidates["final_score"] = (
            0.7 * candidates["base_score"] +
            0.2 * candidates["cat_coverage"] +
            0.1 * candidates["longtail"]
        )
        return candidates

    def _get_fallback_recommendations(self, top_n):
        """ 新用户冷启动策略 """
        return self.videos.nlargest(top_n, "观看次数")

# ==================== 使用示例 ====================
if __name__ == "__main__":
    # 1. 加载数据
    print("正在加载数据...")
    watch_df, videos_df = load_data("user_behavior_details_tfidf.xlsx")
    
    # 2. 初始化推荐引擎
    engine = Recommender(watch_df, videos_df)
    
    # 3. 为用户生成推荐
    user_id = "U0021"
    print(f"\n为用户 {user_id} 生成推荐:")
    recs = engine.get_user_recommendations(user_id, top_n=5)
    print(recs[["所属类别", "观看次数", "final_score"]])
