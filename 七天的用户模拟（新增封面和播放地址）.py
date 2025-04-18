��Ƶ�û���Ϊģ��ϵͳ - �����棨��ÿ�չۿ�ͳ�ƺ��������У�

�����ص㣺
1. ��Excel������Ƶ����
2. �Զ�������Ƶ��TF-IDF+��ξ��ࣩ
3. ģ���û�7����Ϊ
4. ���ɰ���6��������ı��棺
   - ����ͳ��
   - �ۿ���¼��������У�
   - ���޼�¼
   - ��Ƶͳ������
   - ÿ�չۿ�ͳ�ƣ�������
   - ������Ƶ���У�������
"""

# ==================== ���������� ====================
import random
import os
from datetime import datetime, timedelta
from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import AgglomerativeClustering
import numpy as np
from tqdm import tqdm

# ==================== ȫ������ ====================
NUM_USERS = 1000
MIN_INTERESTS = 4
MAX_INTERESTS = 6
OUTPUT_FILENAME = "user_behavior_7days.xlsx"
DAYS_TO_SIMULATE = 7

# ==================== �����ඨ�� ====================
class Video:
    """��Ƶ���󣬴洢Ԫ���ݺ���Ϊͳ��"""
    def __init__(self, url, keywords, cover_url=None, play_url=None):
        self.url = url.strip()
        self.keywords = [kw for kw in (keywords.split() if keywords else []) if kw]
        self.cover_url = cover_url.strip() if cover_url else None  # ���������ַ
        self.play_url = play_url.strip() if play_url else None      # �������ŵ�ַ
        self.category = None
        self.like_count = 0
        self.viewers = set()  # �洢�ۿ������û�ID���Զ�ȥ�أ�
        self.liked_users = set()  # �洢���޹����û�ID

class User:
    """�û����󣬰�����Ȥƫ�ú���Ϊ��¼"""
    def __init__(self, user_id, all_categories):
        self.user_id = f"U{user_id:04d}"
        self.interests = self._generate_interests(all_categories)
        self.watched_videos = []  # ��ʽ: [(url, timestamp_str), ...]
        self.liked_videos = []    # ��ʽ: [(url, timestamp_str), ...]
    
    def _generate_interests(self, categories):
        """��������û���Ȥ�ֲ�"""
        if not categories:
            return {}
        # ȷ����Ȥ����������ʵ�������
        num_interests = random.randint(
            min(MIN_INTERESTS, len(categories)),
            min(MAX_INTERESTS, len(categories)))
        chosen_cats = random.sample(categories, num_interests)
        return {cat: random.uniform(0.7, 1.0) for cat in chosen_cats}

# ==================== ���ݴ����� ====================
def load_excel_data(file_path):
    """��Excel������Ƶ���ݣ��Զ����URL�͹ؼ����У�"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"�����ļ� {file_path} ������")

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        videos = []
        col_mapping = {}

        # �Զ������λ�ã�������Ӣ��������
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and 'url' in header.lower():
                col_mapping['url'] = col
            elif header and ('keyword' in header.lower() or '��ǩ' in header.lower()):
                col_mapping['keywords'] = col
            elif header and ('cover' in header.lower() or '����' in header.lower()):  # ���������ַ���
                col_mapping['cover_url'] = col
            elif header and ('play' in header.lower() or '����' in header.lower()):    # �������ŵ�ַ���
                col_mapping['play_url'] = col

        if not col_mapping.get('url'):
            raise ValueError("δ�ҵ���Ҫ��URL��")

        # ��ȡ��Ƶ����
        for row in range(2, ws.max_row + 1):
            url = ws.cell(row=row, column=col_mapping['url']).value
            keywords = ws.cell(row=row, column=col_mapping['keywords']).value if 'keywords' in col_mapping else ""
            cover_url = ws.cell(row=row, column=col_mapping['cover_url']).value if 'cover_url' in col_mapping else None
            play_url = ws.cell(row=row, column=col_mapping['play_url']).value if 'play_url' in col_mapping else None
            
            if url:
                videos.append(Video(url, keywords, cover_url, play_url))

        return videos
    except Exception as e:
        raise ValueError(f"Excel��ȡ����: {str(e)}")

def auto_categorize(videos):
    """ʹ�ò�ξ������Ƶ���ࣨ���400�ࣩ"""
    if not videos:
        print("����: û����Ƶ���ݿɹ�����")
        return {}

    # ׼���ı�����
    video_docs = [' '.join(v.keywords) for v in videos if v.keywords]
    if not video_docs:
        print("����: ����Ч�ؼ���")
        return {}

    # TF-IDF������ȡ
    tfidf = TfidfVectorizer(
        ngram_range=(1, 2),
        min_df=3,
        max_features=5000,
        token_pattern=r'\b\w+\b'
    )
    try:
        tfidf_matrix = tfidf.fit_transform(video_docs)
    except ValueError:
        print("����: �������㣬��������")
        return {}

    # ��ξ���
    MAX_CATEGORIES = 400
    actual_clusters = min(MAX_CATEGORIES, len(videos))
    
    clustering = AgglomerativeClustering(
        n_clusters=actual_clusters,
        metric='cosine',
        linkage='average'
    )
    clusters = clustering.fit_predict(tfidf_matrix.toarray())

    # ��������ǩ�����ڸ�Ƶ�ؼ��ʣ�
    category_map = {}
    cluster_keywords = defaultdict(list)
    feature_names = tfidf.get_feature_names_out()

    for idx, cluster_id in enumerate(clusters):
        vec = tfidf_matrix[idx]
        top_features = sorted(zip(vec.indices, vec.data), key=lambda x: -x[1])[:3]
        selected_keywords = [feature_names[i] for i, _ in top_features]
        cluster_keywords[cluster_id].extend(selected_keywords)

    for cluster_id, keywords in cluster_keywords.items():
        top_keywords = [kw for kw, _ in Counter(keywords).most_common(3)]
        label = "_".join(top_keywords[:2]) if top_keywords else "����"
        category_map[cluster_id] = f"CAT_{label}"

    # �������
    for idx, video in enumerate(videos):
        if video.keywords:
            video.category = category_map.get(clusters[idx], "����")
        else:
            video.category = "�޹ؼ���"

    # ��ӡ����ͳ��
    print("\n=== ����ͳ�� ===")
    category_counts = Counter(v.category for v in videos)
    for cat, count in category_counts.most_common(15):
        print(f"{cat:<25} {count:<10}")
    print(f"�ܷ�����: {len(category_counts)}")

    return category_map

# ==================== ��Ϊģ�⺯�� ====================
def simulate_behavior(users, videos):
    """ģ��7����û��ۿ��͵�����Ϊ"""
    if not users or not videos:
        print("����: ȱ���û�����Ƶ����")
        return

    # �����������Ƶ
    cat_index = defaultdict(list)
    for v in videos:
        if v.category:
            cat_index[v.category].append(v)

    # 7�����ڷ�Χ����6��ǰ�����죩
    base_date = datetime.now().date()
    date_range = [base_date - timedelta(days=i) for i in range(DAYS_TO_SIMULATE-1, -1, -1)]

    for user in tqdm(users, desc="ģ���û���Ϊ"):
        for day in date_range:
            # ÿ��1-3���Ự
            for _ in range(random.randint(1, 3)):
                # ���ɻỰʱ�䣨����8-23�㣩
                session_time = datetime.combine(day, datetime.min.time()).replace(
                    hour=random.randint(8, 23),
                    minute=random.randint(0, 59))

                # ÿ���Ự�ۿ�10-30����Ƶ
                for _ in range(random.randint(10, 30)):
                    # 90%���ʰ���Ȥѡ����Ƶ
                    if user.interests and random.random() < 0.9:
                        chosen_cat = random.choices(
                            list(user.interests.keys()),
                            weights=list(user.interests.values()),
                            k=1
                        )[0]
                        candidates = cat_index.get(chosen_cat, [])
                    else:
                        candidates = videos

                    if candidates:
                        video = random.choice(candidates)
                        _record_watch(user, video, session_time)

                        # ��̬������޸��ʣ�����20% + ��ȤȨ�أ�
                        like_prob = 0.2
                        if video.category in user.interests:
                            like_prob += user.interests[video.category] * 0.8

                        if random.random() < like_prob:
                            _record_like(user, video, session_time)

def _record_watch(user, video, time):
    """��¼�ۿ���Ϊ���Զ�ȥ�أ�"""
    entry = (video.url, time.strftime('%Y-%m-%d %H:%M:%S'))
    user.watched_videos.append(entry)
    video.viewers.add(user.user_id)

def _record_like(user, video, time):
    """��¼������Ϊ��ȷ�����û����ظ����ޣ�"""
    if user.user_id not in video.liked_users:
        entry = (video.url, time.strftime('%Y-%m-%d %H:%M:%S'))
        user.liked_videos.append(entry)
        video.like_count += 1
        video.liked_users.add(user.user_id)

# ==================== ������ÿ�չۿ�ͳ�� ====================
def count_daily_views(videos, users, days_to_simulate=7):
    """ͳ��ÿ����Ƶÿ��Ĺۿ�����"""
    base_date = datetime.now().date()
    date_range = [(base_date - timedelta(days=i)).strftime('%Y-%m-%d') 
                 for i in range(days_to_simulate-1, -1, -1)]
    
    daily_stats = {v.url: {date: 0 for date in date_range} for v in videos}
    
    for user in users:
        for url, time_str in user.watched_videos:
            watch_date = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
            if watch_date in date_range:
                daily_stats[url][watch_date] += 1
    
    return daily_stats

# ==================== ������������Ƶ���� ====================
def merge_sort_videos(videos, key_func):
    """�鲢����ʵ�֣��ȶ�����"""
    if len(videos) <= 1:
        return videos
    
    mid = len(videos) // 2
    left = merge_sort_videos(videos[:mid], key_func)
    right = merge_sort_videos(videos[mid:], key_func)
    return _merge(left, right, key_func)

def _merge(left, right, key_func):
    """�ϲ���������������б�"""
    result = []
    i = j = 0
    
    while i < len(left) and j < len(right):
        # �������У��ȶȸߵ���ǰ��
        if key_func(left[i]) > key_func(right[j]):
            result.append(left[i])
            i += 1
        else:
            result.append(right[j])
            j += 1
    
    result.extend(left[i:])
    result.extend(right[j:])
    return result

def generate_hot_ranking(videos, top_n=100):
    """����������Ƶ���а�"""
    def sort_key(video):
        # �ȶȹ�ʽ���ۿ�����*0.2 + ������*0.8
        return len(video.viewers) * 0.2 + video.like_count * 0.8
    
    # ʹ�ù鲢����
    sorted_videos = merge_sort_videos(videos, key_func=sort_key)
    
    # ������������
    ranking = []
    for rank, video in enumerate(sorted_videos[:top_n], 1):
        ranking.append({
            "����": rank,
            "��ƵURL": video.url,
            "���": video.category,
            "�ۿ���": len(video.viewers),
            "������": video.like_count,
            "�ȶ�": round(sort_key(video), 2)
        })
    
    return ranking

# ==================== �������ɺ��� ====================
def save_detailed_report(users, videos, filename):
    """���ɰ���6���������Excel����"""
    try:
        wb = Workbook()
        
        # Sheet 1: ����ͳ��
        _create_category_sheet(wb, videos)
        
        # Sheet 2: �ۿ���¼��������У�
        _create_watch_history_sheet(wb, users)
        
        # Sheet 3: ���޼�¼
        _create_like_history_sheet(wb, users)
        
        # Sheet 4: ��Ƶͳ��
        _create_video_stats_sheet(wb, videos)
        
        # Sheet 5: ÿ�չۿ�ͳ�ƣ�������
        _create_daily_views_sheet(wb, videos, users)
        
        # Sheet 6: ������Ƶ���У�������
        _create_hot_ranking_sheet(wb, videos)
        
        wb.save(filename)
    except Exception as e:
        raise IOError(f"Excel����ʧ��: {str(e)}")

# �����������ϸ��������
def _create_category_sheet(wb, videos):
    ws = wb.active
    ws.title = "����ͳ��"
    ws.append(["�������", "��Ƶ����"])
    for cat, count in Counter(v.category for v in videos).most_common():
        ws.append([cat, count])

def _create_watch_history_sheet(wb, users):
    ws = wb.create_sheet("�ۿ���¼")
    date_headers = [(datetime.now().date() - timedelta(days=i)).strftime("%m-%d") 
                   for i in range(DAYS_TO_SIMULATE-1, -1, -1)]
    ws.append(["�û�ID"] + date_headers)
    
    for user in users:
        daily_watches = {day: [] for day in date_headers}
        for url, time_str in user.watched_videos:
            day_key = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S').strftime("%m-%d")
            if day_key in daily_watches:
                daily_watches[day_key].append(url)
        
        ws.append([user.user_id] + [", ".join(daily_watches[day]) for day in date_headers])

def _create_like_history_sheet(wb, users):
    ws = wb.create_sheet("���޼�¼")
    ws.append(["�û�ID", "��ƵURL", "����ʱ��"])
    for user in users:
        for url, time in user.liked_videos:
            ws.append([user.user_id, url, time])

def _create_video_stats_sheet(wb, videos):
    ws = wb.create_sheet("��Ƶͳ��")
    # ������ֶε���ͷ
    ws.append(["��ƵURL", "���ŵ�ַ", "�����ַ", "���", "�ۿ�����", "������"])
    for video in videos:
        # ������ֶε�������
        ws.append([
            video.url, 
            video.play_url or "", 
            video.cover_url or "", 
            video.category, 
            len(video.viewers), 
            video.like_count
        ])

def _create_daily_views_sheet(wb, videos, users):
    ws = wb.create_sheet("ÿ�չۿ�")
    daily_views = count_daily_views(videos, users)
    
    dates = [(datetime.now().date() - timedelta(days=i)).strftime("%Y-%m-%d") 
            for i in range(DAYS_TO_SIMULATE-1, -1, -1)]
    ws.append(["��ƵURL"] + dates)
    
    for video in videos:
        ws.append([video.url] + [daily_views[video.url][date] for date in dates])


def _create_hot_ranking_sheet(wb, videos):
    ws = wb.create_sheet("��������")
    ranking = generate_hot_ranking(videos)
    
    # ������ֶε���ͷ
    headers = ["����", "��ƵURL", "���ŵ�ַ", "�����ַ", "���", "�ۿ���", "������", "�ȶ�"]
    ws.append(headers)
    
    for item in ranking:
        video = next((v for v in videos if v.url == item["��ƵURL"]), None)
        # ������ֶε�������
        ws.append([
            item["����"],
            item["��ƵURL"],
            video.play_url if video else "",
            video.cover_url if video else "",
            item["���"],
            item["�ۿ���"],
            item["������"],
            item["�ȶ�"]
        ])

# ==================== ������ ====================
if __name__ == "__main__":
    try:
        print("���ڼ�����Ƶ����...")
        videos = load_excel_data("videos.xlsx")
        print(f"���سɹ����� {len(videos)} ����Ƶ")

        print("���ڽ�����Ƶ����...")
        categories = auto_categorize(videos)

        print(f"���� {NUM_USERS} ��ģ���û�...")
        valid_categories = [c for c in categories.values() if c not in ["����", "�޹ؼ���"]]
        users = [User(i, valid_categories) for i in range(NUM_USERS)]

        print(f"ģ�� {DAYS_TO_SIMULATE} ���û���Ϊ...")
        simulate_behavior(users, videos)

        print(f"���ɱ��� {OUTPUT_FILENAME}...")
        save_detailed_report(users, videos, OUTPUT_FILENAME)

        # ��ӡ����ͳ��
        total_views = sum(len(u.watched_videos) for u in users)
        total_likes = sum(len(u.liked_videos) for u in users)
        print(f"\n=== ģ���� ===")
        print(f"�ܹۿ�����: {total_views}")
        print(f"�ܵ��޴���: {total_likes}")
        print(f"������: {total_likes/total_views:.1%}")
        print(f"�ѱ��浽 {OUTPUT_FILENAME}")

    except Exception as e:
        print(f"������: {str(e)}")