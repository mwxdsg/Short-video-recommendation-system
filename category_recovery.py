# -*- coding: utf-8 -*-
"""
分类恢复模块
用于从现有报告文件中恢复视频分类信息
"""

import os
from openpyxl import load_workbook
from collections import defaultdict


def recover_categories(videos, report_files):
    """
    尝试从现有报告文件中恢复视频分类信息

    参数:
    - videos: 视频对象列表
    - report_files: 可能包含分类信息的报告文件列表（按优先级排序）

    返回:
    - categories: 恢复的分类列表
    - applied_count: 成功应用分类的视频数量
    """
    if not videos:
        return [], 0

    # 创建URL到视频对象的映射
    url_to_video = {v.url: v for v in videos}
    category_info = {}  # URL到分类的映射
    all_categories = set()  # 所有发现的分类

    # 遍历所有候选报告文件
    for report_file in report_files:
        if not os.path.exists(report_file):
            continue

        try:
            print(f"尝试从 {report_file} 恢复分类信息...")
            wb = load_workbook(report_file, read_only=True)

            # 检查每个工作表
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 尝试找出URL列和分类列
                url_col = None
                cat_col = None

                # 读取第一行以定位列
                headers = []
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = row
                    break

                # 查找URL和分类列
                for i, header in enumerate(headers):
                    if header and isinstance(header, str):
                        if 'url' in header.lower():
                            url_col = i
                        elif ('类别' in header.lower() or 'category' in header.lower()
                              or 'cat' in header.lower()):
                            cat_col = i

                # 如果找到了URL列和分类列，则提取信息
                if url_col is not None and cat_col is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if len(row) > max(url_col, cat_col):
                            url = row[url_col]
                            category = row[cat_col]

                            if url and category and url in url_to_video:
                                category_info[url] = category
                                all_categories.add(category)

            print(f"从 {report_file} 找到 {len(category_info)} 个视频的分类信息")

            # 如果已经找到足够多的分类信息，可以提前退出
            if len(category_info) >= len(videos) * 0.7:  # 如果已覆盖70%以上视频
                break

        except Exception as e:
            print(f"从 {report_file} 恢复分类时出错: {str(e)}")
            continue

    # 应用恢复的分类信息到视频对象
    applied_count = 0
    for url, category in category_info.items():
        if url in url_to_video:
            url_to_video[url].category = category
            applied_count += 1

    return list(all_categories), applied_count


def extract_categories_from_video_clusters(file_path):
    """
    从视频聚类报告中提取分类信息
    (这是一个更专用的函数，用于处理视频聚类报告特有的格式)

    参数:
    - file_path: 视频聚类报告文件路径

    返回:
    - 视频URL到分类的映射字典
    """
    if not os.path.exists(file_path):
        return {}

    url_to_category = {}

    try:
        wb = load_workbook(file_path, read_only=True)

        # 首先检查汇总工作表
        if "聚类汇总" in wb.sheetnames:
            # 汇总表通常只有聚类ID和统计信息，不包含单个视频
            pass

        # 然后检查各个聚类工作表
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith("聚类"):
                ws = wb[sheet_name]

                # 找出URL列和类别列
                url_col = None
                cat_col = None

                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    for i, cell in enumerate(row):
                        if cell and isinstance(cell, str):
                            if 'url' in cell.lower():
                                url_col = i
                            elif '类别' in cell.lower() or 'category' in cell.lower():
                                cat_col = i
                    break

                if url_col is not None and cat_col is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if len(row) > max(url_col, cat_col):
                            url = row[url_col]
                            category = row[cat_col]

                            if url and category:
                                url_to_category[url] = category

        return url_to_category

    except Exception as e:
        print(f"从视频聚类报告提取分类时出错: {str(e)}")
        return {}