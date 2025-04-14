# -*- coding: utf-8 -*-

#��Ƶ�û���Ϊģ��ϵͳ - ������

#�����ص㣺
#1. ��Excel������Ƶ���ݣ�����URL�͹ؼ��ʣ�
#2. ʹ��TF-IDF�Ͳ�ξ������Ƶ�Զ����ࣨ���400�ࣩ
#3. ����ģ���û�������Ȥƫ��
#4. ģ��7����û��ۿ��͵�����Ϊ
#5. ������ϸ���棬�ۿ���¼��7����д洢


# ==================== Import Dependencies ====================
import random
import os
from datetime import datetime, timedelta
from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import AgglomerativeClustering
import numpy as np
from tqdm import tqdm

# ==================== Global Configuration ====================
NUM_USERS = 1000  # Number of simulated users
MIN_INTERESTS = 4  # Minimum number of interests per user
MAX_INTERESTS = 6  # Maximum number of interests per user
OUTPUT_FILENAME = "user_behavior_7days.xlsx"  # Output filename
DAYS_TO_SIMULATE = 7  # Number of days to simulate

# ==================== Core Class Definitions ====================
class Video:
    #��Ƶ���󣬰���Ԫ���ݺ���Ϊͳ��
    def __init__(self, url, keywords):
        self.url = url.strip()
        self.keywords = [kw for kw in (keywords.split() if keywords else []) if kw]
        self.category = None
        self.like_count = 0
        self.viewers = set()
        self.liked_users = set()

class User:
    #�û����󣬰�����Ȥƫ�ú���Ϊ��¼
    def __init__(self, user_id, all_categories):
        self.user_id = f"U{user_id:04d}"
        self.interests = self._generate_interests(all_categories)
        self.watched_videos = []  # Format: [(url, timestamp_str), ...]
        self.liked_videos = []    # Format: [(url, timestamp_str), ...]
    
    def _generate_interests(self, categories):
        #��������û���Ȥ�ֲ�
        if not categories:
            return {}
        num_interests = random.randint(
            min(MIN_INTERESTS, len(categories)),
            min(MAX_INTERESTS, len(categories))
        )
        chosen_cats = random.sample(categories, num_interests)
        return {cat: random.uniform(0.7, 1.0) for cat in chosen_cats}

# ==================== Data Processing Functions ====================
def load_excel_data(file_path):
    #��Excel������Ƶ����
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Input file {file_path} not found")

    try:
        wb = load_workbook(file_path)
        ws = wb.active
        videos = []

        # �Զ������λ��
        col_mapping = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and 'url' in header.lower():
                col_mapping['url'] = col
            elif header and ('keyword' in header.lower() or 'tag' in header.lower()):
                col_mapping['keywords'] = col

        if not col_mapping:
            raise ValueError("Required columns (url/keywords) not found in Excel file")

        # ���ж�ȡ��Ƶ����
        for row in range(2, ws.max_row + 1):
            url = ws.cell(row=row, column=col_mapping['url']).value
            keywords = ws.cell(row=row, column=col_mapping['keywords']).value or ""
            if url:
                videos.append(Video(url, keywords))

        return videos
    except Exception as e:
        raise ValueError(f"Excel file reading error: {str(e)}")

def auto_categorize(videos):
    #ʹ�ò�ξ������Ƶ���ࣨ���400�ࣩ
    if not videos:
        print("Warning: No video data available for categorization")
        return {}

    # ������Ƶ�����ĵ�
    video_docs = [' '.join(v.keywords) for v in videos if v.keywords]
    if not video_docs:
        print("Warning: All videos have no valid keywords")
        return {}

    # ����TF-IDF����
    tfidf = TfidfVectorizer(
        ngram_range=(1, 2),
        min_df=3,
        max_features=5000,
        token_pattern=r'\b\w+\b'
    )
    try:
        tfidf_matrix = tfidf.fit_transform(video_docs)
    except ValueError:
        print("Warning: Insufficient features for categorization")
        return {}

    # ��ξ���
    MAX_CATEGORIES = 400
    actual_clusters = min(MAX_CATEGORIES, len(videos))
    
    clustering = AgglomerativeClustering(
        n_clusters=actual_clusters,
        metric='cosine',
        linkage='average'
    )
    clusters = clustering.fit_predict(tfidf_matrix.toarray())

    # ��������ǩ
    category_map = {}
    cluster_keywords = defaultdict(list)
    feature_names = tfidf.get_feature_names_out()

    for idx, cluster_id in enumerate(clusters):
        vec = tfidf_matrix[idx]
        feature_weights = zip(vec.indices, vec.data)
        sorted_features = sorted(feature_weights, key=lambda x: -x[1])[:3]
        selected_keywords = [feature_names[i] for i, _ in sorted_features]
        cluster_keywords[cluster_id].extend(selected_keywords)

    for cluster_id, keywords in cluster_keywords.items():
        counter = Counter(keywords)
        top_keywords = [kw for kw, _ in counter.most_common(3)]
        label = "_".join(top_keywords[:2]) if top_keywords else "Other"
        category_map[cluster_id] = f"CAT_{label}"

    # ������Ƶ���
    category_counter = Counter()
    for idx, video in enumerate(videos):
        if video.keywords:
            cluster_id = clusters[idx]
            video.category = category_map.get(cluster_id, "Other")
            category_counter[video.category] += 1
        else:
            video.category = "No_Keywords"
            category_counter[video.category] += 1

    # ��ӡ����ͳ��
    print("\n=== Category Statistics ===")
    print("{:<25} {:<10}".format("Category Name", "Video Count"))
    print("-" * 45)
    for cat, count in category_counter.most_common(15):
        print("{:<25} {:<10}".format(cat, count))
    print(f"(Total {len(category_counter)} categories)")

    return category_map

# ==================== Behavior Simulation Functions ====================
def simulate_behavior(users, videos):
    #ģ��7����û��ۿ��͵�����Ϊ
    if not users or not videos:
        print("Warning: Missing user or video data")
        return

    # ���������Ƶ����
    cat_index = defaultdict(list)
    for v in videos:
        if v.category:
            cat_index[v.category].append(v)

    # ����7�����ڷ�Χ����6��ǰ�����죩
    base_date = datetime.now().date()
    date_range = [base_date - timedelta(days=i) for i in range(DAYS_TO_SIMULATE-1, -1, -1)]

    for user in tqdm(users, desc="Simulating user behavior"):
        for day in date_range:
            # ÿ��1-3���Ự
            for _ in range(random.randint(1, 3)):
                # ���ɻỰʱ�䣨����8-23�㣩
                session_time = datetime.combine(day, datetime.min.time()).replace(
                    hour=random.randint(8, 23),
                    minute=random.randint(0, 59))

                # ÿ���Ự�ۿ�10-30����Ƶ
                for _ in range(random.randint(10, 30)):
                    # ����Ȥѡ����Ƶ
                    if user.interests and random.random() < 0.9:
                        chosen_cat = random.choices(
                            list(user.interests.keys()),
                            weights=list(user.interests.values()),
                            k=1
                        )[0]
                        candidates = cat_index.get(chosen_cat, [])
                    else:
                        candidates = videos

                    if candidates:
                        video = random.choice(candidates)
                        _record_watch(user, video, session_time)

                        # ������޸���
                        like_prob = 0.2
                        if video.category in user.interests:
                            like_prob += 0.8 * user.interests[video.category]

                        if random.random() < like_prob:
                            _record_like(user, video, session_time)

def _record_watch(user, video, time):
    #��¼�ۿ���Ϊ
    entry = (video.url, time.strftime('%Y-%m-%d %H:%M:%S'))
    user.watched_videos.append(entry)
    video.viewers.add(user.user_id)

def _record_like(user, video, time):
    #��¼������Ϊ
    if user.user_id not in video.liked_users:
        entry = (video.url, time.strftime('%Y-%m-%d %H:%M:%S'))
        user.liked_videos.append(entry)
        video.like_count += 1
        video.liked_users.add(user.user_id)

# ==================== Report Generation Functions ====================
def save_detailed_report(users, videos, filename):
    #����Excel���棨�ۿ���¼��7����У�
    try:
        wb = Workbook()

        # Sheet 1: Category Statistics
        ws_cats = wb.active
        ws_cats.title = "Category Stats"
        ws_cats.append(["Category Name", "Video Count"])
        cat_stats = defaultdict(int)
        for v in videos:
            if v.category:
                cat_stats[v.category] += 1
        for cat, count in sorted(cat_stats.items(), key=lambda x: x[1], reverse=True):
            ws_cats.append([cat, count])

        # Sheet 2: Watch History (by day)
        ws_watches = wb.create_sheet("Watch History")
        
        # ����7����б���
        base_date = datetime.now().date()
        date_headers = [(base_date - timedelta(days=i)).strftime("%m-%d") 
                       for i in range(DAYS_TO_SIMULATE-1, -1, -1)]
        
        headers = ["User ID"] + date_headers
        ws_watches.append(headers)

        # Ϊÿ���û��ռ�7��Ĺۿ���¼
        for user in users:
            daily_watches = {day: [] for day in date_headers}
            
            for url, time_str in user.watched_videos:
                watch_time = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
                day_key = watch_time.strftime("%m-%d")
                if day_key in daily_watches:
                    daily_watches[day_key].append(url)
            
            # д��������
            row = [user.user_id]
            for day in date_headers:
                row.append(", ".join(daily_watches[day]))
            ws_watches.append(row)

        # Sheet 3: Like History
        ws_likes = wb.create_sheet("Like History")
        ws_likes.append(["User ID", "Video URL", "Like Time"])
        for user in users:
            for url, time in user.liked_videos:
                ws_likes.append([user.user_id, url, time])

        # Sheet 4: Video Statistics
        ws_videos = wb.create_sheet("Video Statistics")
        ws_videos.append(["Video URL", "Category", "View Count", "Like Count"])
        for video in videos:
            view_count = len(video.viewers)
            ws_videos.append([video.url, video.category, view_count, video.like_count])

        wb.save(filename)
    except Exception as e:
        raise IOError(f"Failed to save Excel file: {str(e)}")

# ==================== Main Program ====================
if __name__ == "__main__":
    try:
        # Step 1: Load video data
        print("Loading video data...")
        videos = load_excel_data("videos.xlsx")
        print(f"Successfully loaded {len(videos)} videos")

        # Step 2: Auto-categorize videos
        print("Categorizing videos using TF-IDF...")
        cat_names = auto_categorize(videos)

        # Step 3: Generate simulated users
        print(f"Creating {NUM_USERS} simulated users...")
        valid_categories = [cat for cat in cat_names.values() if cat not in ["Other", "No_Keywords"]]
        users = [User(i, valid_categories) for i in range(NUM_USERS)]
        
        # Step 4: Simulate 7-day user behavior
        print(f"Simulating {DAYS_TO_SIMULATE} days of user behavior...")
        simulate_behavior(users, videos)

        # Step 5: Save report
        print(f"Generating report file {OUTPUT_FILENAME}...")
        save_detailed_report(users, videos, OUTPUT_FILENAME)

        # Summary statistics
        total_views = sum(len(u.watched_videos) for u in users)
        total_likes = sum(len(u.liked_videos) for u in users)

        print("\n=== Simulation Summary ===")
        print(f"Total views: {total_views}")
        print(f"Total likes: {total_likes}")
        print(f"Like rate: {total_likes / total_views:.1%}")
        print(f"Report saved to {OUTPUT_FILENAME}")

    except FileNotFoundError as e:
        print(f"File error: {str(e)}")
    except ValueError as e:
        print(f"Data error: {str(e)}")
    except Exception as e:
        print(f"Unexpected error: {str(e)}")